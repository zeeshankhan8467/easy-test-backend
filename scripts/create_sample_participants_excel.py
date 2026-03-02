#!/usr/bin/env python3
"""
Generate sample_participants_import.xlsx for testing participant import.
Run from backend: cd easy-test-backend && source venv/bin/activate && python scripts/create_sample_participants_excel.py
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Activate venv: cd easy-test-backend && source venv/bin/activate")
    sys.exit(1)

HEADERS = [
    "Name", "Clicker ID", "Roll No.", "Admission No.", "Class", "Subject",
    "Section", "Team", "Group", "House", "Gender", "City", "UID",
    "Employee Code", "Teacher Name", "Email ID",
]

ROWS = [
    ["Rahul Sharma", "S001", "1", "ADM2024001", "6", "Math", "A", "Alpha", "G1", "Red", "Male", "Mumbai", "UID001", "", "Mr. Kumar", "rahul.sharma@example.com"],
    ["Priya Patel", "S002", "2", "ADM2024002", "6", "Science", "A", "Alpha", "G1", "Blue", "Female", "Delhi", "UID002", "", "Mrs. Sharma", "priya.patel@example.com"],
    ["Amit Singh", "S003", "3", "ADM2024003", "7", "English", "B", "Beta", "G2", "Green", "Male", "Bangalore", "UID003", "", "Mr. Patel", "amit.singh@example.com"],
    ["Sneha Reddy", "S004", "4", "ADM2024004", "7", "History", "B", "Beta", "G2", "Yellow", "Female", "Chennai", "UID004", "", "Ms. Reddy", "sneha.reddy@example.com"],
    ["Vikram Desai", "S005", "5", "ADM2024005", "8", "Geography", "A", "Gamma", "G3", "Red", "Male", "Hyderabad", "UID005", "EMP005", "Mr. Kumar", "vikram.desai@example.com"],
    ["Anita Nair", "S006", "6", "ADM2024006", "8", "Math", "A", "Gamma", "G3", "Blue", "Female", "Pune", "UID006", "", "Mrs. Sharma", "anita.nair@example.com"],
    ["Rohan Mehta", "S007", "7", "ADM2024007", "9", "Physics", "C", "Delta", "G4", "Green", "Male", "Kolkata", "UID007", "", "Mr. Patel", "rohan.mehta@example.com"],
    ["Kavita Joshi", "S008", "8", "ADM2024008", "9", "Chemistry", "C", "Delta", "G4", "Yellow", "Female", "Ahmedabad", "UID008", "", "Ms. Reddy", "kavita.joshi@example.com"],
    ["Arjun Iyer", "S009", "9", "ADM2024009", "10", "Computer", "A", "Alpha", "G1", "Red", "Male", "Jaipur", "UID009", "", "Mr. Kumar", "arjun.iyer@example.com"],
    ["Divya Menon", "S010", "10", "ADM2024010", "10", "Economics", "A", "Alpha", "G1", "Blue", "Female", "Lucknow", "UID010", "", "Mrs. Sharma", "divya.menon@example.com"],
]


def main():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_path = os.path.join(base_dir, "sample_participants_import.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row_data in enumerate(ROWS, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    col_letters = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]
    for letter in col_letters:
        ws.column_dimensions[letter].width = 16

    wb.save(out_path)
    print(f"Created: {out_path}")
    print(f"  Rows: 1 header + {len(ROWS)} sample participants")
    print("  Use in EasyTest: Participants → Import CSV/Excel")


if __name__ == "__main__":
    main()
