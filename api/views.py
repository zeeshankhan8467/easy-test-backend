from rest_framework import viewsets, status, renderers, serializers as drf_serializers
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken


class BinaryFileRenderer(renderers.BaseRenderer):
    """Pass bytes through unchanged for file download responses."""
    media_type = 'application/octet-stream'
    format = 'binary'

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data if isinstance(data, bytes) else b''
from django.db.models import Q, Count, Avg, Max, Min, Sum
from django.db.models.functions import Length
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_GET
from django.core.mail import EmailMessage, get_connection
from rest_framework_simplejwt.authentication import JWTAuthentication
from datetime import timedelta
import pandas as pd
import json
import logging
import re
import os
import uuid
import html as html_module
from urllib.parse import quote
from io import BytesIO
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

logger = logging.getLogger(__name__)

from .models import (
    Exam, Question, ExamQuestion, Participant,
    ExamParticipant, ExamAttempt, Answer, School, UserProfile,
    DailyAttendance,
    ROLE_SUPER_ADMIN, ROLE_SCHOOL_ADMIN, ROLE_TEACHER,
)
from .permissions import (
    scope_exams_queryset, scope_participants_queryset, scope_schools_queryset,
    can_create_school_admin, can_create_teacher, get_user_school_id, get_user_role,
)
from .serializers import (
    UserSerializer, LoginSerializer, ExamSerializer, ExamCreateUpdateSerializer,
    QuestionSerializer, ParticipantSerializer, ExamParticipantSerializer,
    ExamAttemptSerializer, QuestionAnalysisSerializer, ParticipantResultSerializer,
    ExamReportSerializer, DashboardStatsSerializer, RecentExamSerializer,
    PerformanceDataSerializer, DashboardDataSerializer, LeaderboardEntrySerializer,
    LeaderboardSerializer, SchoolSerializer, CreateSchoolAdminSerializer, CreateTeacherSerializer,
)


# Authentication Views
@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    serializer = LoginSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.validated_data['user']
        refresh = RefreshToken.for_user(user)
        
        user_data = UserSerializer(user).data
        
        return Response({
            'token': str(refresh.access_token),
            'user': user_data
        })
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Exam Views
class ExamViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = Exam.objects.all().prefetch_related('exam_questions__question').select_related('school', 'created_by')
        qs = scope_exams_queryset(qs, self.request.user)
        role = get_user_role(self.request.user)
        params = self.request.query_params
        if role == ROLE_SUPER_ADMIN:
            school_id = params.get('school_id')
            if school_id is not None and school_id != '':
                try:
                    qs = qs.filter(school_id=int(school_id))
                except (ValueError, TypeError):
                    pass
            owner_user_id = params.get('owner_user_id')
            if owner_user_id is not None and owner_user_id != '':
                try:
                    qs = qs.filter(created_by_id=int(owner_user_id))
                except (ValueError, TypeError):
                    pass
        elif role == ROLE_SCHOOL_ADMIN:
            owner_user_id = params.get('owner_user_id')
            if owner_user_id is not None and owner_user_id != '':
                try:
                    qs = qs.filter(created_by_id=int(owner_user_id))
                except (ValueError, TypeError):
                    pass
        return qs
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ExamCreateUpdateSerializer
        return ExamSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Allow creating exam without questions (draft mode)
        # Questions can be added later before freezing
        exam = serializer.save()
        return Response(ExamSerializer(exam).data, status=status.HTTP_201_CREATED)
    
    def update(self, request, *args, **kwargs):
        exam = self.get_object()
        
        # Check if exam can be edited
        if exam.status == 'frozen':
            return Response(
                {'error': 'Cannot edit a frozen exam'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = self.get_serializer(exam, data=request.data, partial=kwargs.get('partial', False))
        serializer.is_valid(raise_exception=True)
        
        # Allow removing all questions (draft mode)
        # But validate before freezing
        exam = serializer.save()
        return Response(ExamSerializer(exam).data)

    @action(detail=True, methods=['post'])
    def freeze(self, request, pk=None):
        exam = self.get_object()
        
        # Validate exam can be frozen
        if exam.status == 'frozen':
            return Response(
                {'error': 'Exam is already frozen'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if exam.exam_questions.count() == 0:
            return Response(
                {'error': 'Cannot freeze an exam without questions'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generate snapshot
        import hashlib
        import json as json_lib
        
        option_display = (request.data.get('option_display') or request.data.get('optionDisplay') or 'alpha')
        if isinstance(option_display, str):
            option_display = option_display.strip().lower()
        if option_display not in ('alpha', 'numeric'):
            option_display = 'alpha'
        
        questions = exam.exam_questions.select_related('question').order_by('order')
        snapshot_data = {
            'exam_id': exam.id,
            'title': exam.title,
            'description': exam.description,
            'duration': exam.duration,
            'revisable': exam.revisable,
            'show_live_response': getattr(exam, 'show_live_response', False),
            'show_response_after_completion': getattr(exam, 'show_response_after_completion', True),
            'question_change_automatic': getattr(exam, 'question_change_automatic', False),
            'option_display': option_display,
            'frozen_at': timezone.now().isoformat(),
            'questions': []
        }
        
        for eq in questions:
            q_display = getattr(eq.question, 'option_display', None) or option_display
            if str(q_display).lower() not in ('alpha', 'numeric'):
                q_display = option_display
            snapshot_data['questions'].append({
                'question_id': eq.question.id,
                'order': eq.order,
                'text': eq.question.text,
                'type': eq.question.type,
                'options': eq.question.options,
                'correct_answer': eq.question.correct_answer,
                'difficulty': eq.question.difficulty,
                'option_display': q_display,
                'positive_marks': float(eq.positive_marks),
                'negative_marks': float(eq.negative_marks),
                'is_optional': eq.is_optional,
            })
        
        # Generate version/checksum
        snapshot_json = json_lib.dumps(snapshot_data, sort_keys=True)
        snapshot_version = hashlib.md5(snapshot_json.encode()).hexdigest()
        
        # Freeze exam
        exam.status = 'frozen'
        exam.frozen = True
        exam.snapshot_data = snapshot_data
        exam.snapshot_version = snapshot_version
        exam.save()
        
        return Response(ExamSerializer(exam).data)

    @action(detail=True, methods=['get'], url_path='export')
    def export_report_action(self, request, pk=None):
        """Export exam report as Excel or CSV. Query param: format=excel|csv. Uses same access as exam_report (any authenticated user)."""
        try:
            exam = Exam.objects.get(pk=pk)
        except Exam.DoesNotExist:
            return Response({'error': 'Exam not found'}, status=status.HTTP_404_NOT_FOUND)
        return _build_export_http_response(request, exam)

    @action(detail=False, methods=['get'])
    def available_questions(self, request):
        """Get available questions for selection (not already in exam). Scoped by role."""
        exam_id = request.query_params.get('exam_id')
        difficulty = request.query_params.get('difficulty')
        qtype = request.query_params.get('type')
        search = request.query_params.get('search', '')

        # Scope by role: Super Admin all, School Admin same school, Teacher own only
        user = request.user
        role = get_user_role(user)
        if role == ROLE_SUPER_ADMIN:
            queryset = Question.objects.all()
        elif role == ROLE_SCHOOL_ADMIN:
            school_id = get_user_school_id(user)
            if not school_id:
                queryset = Question.objects.none()
            else:
                queryset = Question.objects.filter(created_by__profile__school_id=school_id)
        elif role == ROLE_TEACHER:
            queryset = Question.objects.filter(created_by=user)
        else:
            queryset = Question.objects.filter(created_by=user)

        # Filter by difficulty
        if difficulty and difficulty != 'all':
            queryset = queryset.filter(difficulty=difficulty)
        
        # Filter by type
        if qtype and qtype != 'all':
            queryset = queryset.filter(type=qtype)
        
        # Search in text
        if search:
            queryset = queryset.filter(text__icontains=search)
        
        # Exclude questions already in exam (if editing)
        if exam_id:
            try:
                exam = Exam.objects.get(id=exam_id, created_by=request.user)
                existing_question_ids = exam.exam_questions.values_list('question_id', flat=True)
                queryset = queryset.exclude(id__in=existing_question_ids)
            except Exam.DoesNotExist:
                pass
        
        serializer = QuestionSerializer(queryset[:100], many=True)  # Limit to 100
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def snapshot(self, request, pk=None):
        exam = self.get_object()
        
        # Return stored snapshot if frozen, otherwise generate current snapshot
        if exam.snapshot_data:
            snapshot_data = exam.snapshot_data
            snapshot_data.setdefault('option_display', 'alpha')
            snapshot_data.setdefault('show_live_response', False)
            snapshot_data.setdefault('show_response_after_completion', True)
            snapshot_data.setdefault('question_change_automatic', False)
        else:
            questions = exam.exam_questions.select_related('question').order_by('order')
            snapshot_data = {
                'exam_id': exam.id,
                'title': exam.title,
                'description': exam.description,
                'duration': exam.duration,
                'revisable': exam.revisable,
                'show_live_response': getattr(exam, 'show_live_response', False),
                'show_response_after_completion': getattr(exam, 'show_response_after_completion', True),
            'question_change_automatic': getattr(exam, 'question_change_automatic', False),
                'option_display': 'alpha',
                'generated_at': timezone.now().isoformat(),
                'questions': []
            }
            
            for eq in questions:
                q_display = getattr(eq.question, 'option_display', 'alpha')
                if str(q_display).lower() not in ('alpha', 'numeric'):
                    q_display = 'alpha'
                snapshot_data['questions'].append({
                    'question_id': eq.question.id,
                    'order': eq.order,
                    'text': eq.question.text,
                    'type': eq.question.type,
                    'options': eq.question.options,
                    'correct_answer': eq.question.correct_answer,
                    'difficulty': eq.question.difficulty,
                    'option_display': q_display,
                    'positive_marks': float(eq.positive_marks),
                    'negative_marks': float(eq.negative_marks),
                    'is_optional': eq.is_optional,
                })
        
        # Include version for client validation (only when frozen)
        if exam.snapshot_version:
            snapshot_data['snapshot_version'] = exam.snapshot_version
        response = Response(snapshot_data)
        response['Content-Disposition'] = f'attachment; filename="exam-{exam.id}-snapshot.json"'
        return response

    @action(detail=True, methods=['post'])
    def sync_live_results(self, request, pk=None):
        """
        Submit live clicker responses and attendance from EasyTest Live app.
        Body: {
            "responses": [
                {"participant_id": 1, "question_id": 2, "selected_answer": 0, "answered_at": "ISO8601"},
                {"clicker_id": "123", ...}  // alternative: resolve participant by clicker_id
            ],
            "attendance": [1, 2, 3]  // optional: participant_ids to mark present
        }
        selected_answer: 0-based index (MCQ) or list of indices (multiple_select). Letter "A"=0, "B"=1, etc.
        """
        exam = self.get_object()
        if exam.status not in ('frozen', 'completed'):
            logger.warning('[sync_live_results] Exam %s not frozen/completed, status=%s', pk, exam.status)
            return Response(
                {'error': 'Exam must be frozen to accept live results'},
                status=status.HTTP_400_BAD_REQUEST
            )

        responses_data = request.data.get('responses', [])
        attendance_ids = request.data.get('attendance', [])
        exam_started_at_raw = request.data.get('exam_started_at')
        exam_started_at_parsed = None
        if exam_started_at_raw:
            try:
                from datetime import datetime
                dt = datetime.fromisoformat(str(exam_started_at_raw).replace('Z', '+00:00'))
                if timezone.is_naive(dt):
                    dt = timezone.make_aware(dt)
                exam_started_at_parsed = dt
            except (ValueError, TypeError):
                pass
        logger.info(
            '[sync_live_results] TIME_TAKEN DEBUG: exam_started_at raw=%s parsed=%s',
            exam_started_at_raw, exam_started_at_parsed
        )
        logger.info(
            '[sync_live_results] Exam id=%s: received %d responses, %d attendance',
            pk, len(responses_data), len(attendance_ids)
        )

        # Build snapshot question lookup (question_id -> correct_answer, positive_marks, negative_marks)
        snapshot = exam.snapshot_data or {}
        questions_list = snapshot.get('questions', [])
        if questions_list:
            snapshot_questions = {int(q['question_id']): q for q in questions_list}
        else:
            snapshot_questions = {}
            for eq in exam.exam_questions.select_related('question').order_by('order'):
                snapshot_questions[eq.question_id] = {
                    'correct_answer': eq.question.correct_answer,
                    'positive_marks': float(eq.positive_marks),
                    'negative_marks': float(eq.negative_marks),
                }

        # Resolve participant by id or clicker_id; create one from clicker_id if missing (e.g. deviceId fallback when SDK keySN is empty)
        # clicker_id is unique per exam author (teacher), not globally — scope lookups to this exam / exam.created_by.
        def _participant_by_clicker_for_exam(clicker_id_str):
            if not clicker_id_str:
                return None
            p = (
                Participant.objects.filter(
                    clicker_id=clicker_id_str,
                    participant_exams__exam=exam,
                )
                .first()
            )
            if p:
                return p
            if exam.created_by_id:
                p = (
                    Participant.objects.filter(
                        clicker_id=clicker_id_str,
                        created_by_id=exam.created_by_id,
                    )
                    .first()
                )
                if p:
                    return p
            # Legacy / edge case: fall back to global match (e.g. old rows without created_by)
            return Participant.objects.filter(clicker_id=clicker_id_str).first()

        def get_or_create_participant_for_clicker(participant_id=None, clicker_id=None):
            if participant_id:
                try:
                    p = Participant.objects.get(id=participant_id)
                    ExamParticipant.objects.get_or_create(exam=exam, participant=p)
                    return p
                except Participant.DoesNotExist:
                    pass
            if clicker_id:
                clicker_id_str = str(clicker_id).strip()
                if not clicker_id_str:
                    return None
                p = _participant_by_clicker_for_exam(clicker_id_str)
                if p:
                    ExamParticipant.objects.get_or_create(exam=exam, participant=p)
                    return p
                # When app sends deviceId fallback (d1_timestamp) because SDK keySN is empty, match to participant with clicker_id = number (e.g. "1")
                if clicker_id_str.startswith('d') and '_' in clicker_id_str:
                    num_part = clicker_id_str[1:].split('_')[0]
                    if num_part.isdigit():
                        p = _participant_by_clicker_for_exam(num_part)
                        if p:
                            ExamParticipant.objects.get_or_create(exam=exam, participant=p)
                            return p
                # Auto-create participant only if no match (e.g. when SDK sends empty keySN and no participant has that clicker number)
                safe_id = ''.join(c if c.isalnum() or c in '_-' else '_' for c in clicker_id_str)[:50]
                uid = exam.created_by_id or 0
                email = f'clicker-e{exam.id}-u{uid}-{safe_id}@easytest.local'
                defaults = {
                    'name': 'Student',
                    'email': email,
                }
                if exam.school_id:
                    defaults['school_id'] = exam.school_id
                if exam.created_by_id:
                    p, created = Participant.objects.get_or_create(
                        clicker_id=clicker_id_str,
                        created_by_id=exam.created_by_id,
                        defaults=defaults,
                    )
                else:
                    p, created = Participant.objects.get_or_create(
                        clicker_id=clicker_id_str,
                        defaults=defaults,
                    )
                if created:
                    logger.info('[sync_live_results] Auto-created participant id=%s clicker_id=%s for exam %s', p.id, clicker_id_str, pk)
                ExamParticipant.objects.get_or_create(exam=exam, participant=p)
                return p
            return None

        # Normalize selected_answer to 0-based index or list
        def normalize_selected(val):
            if val is None:
                return None
            if isinstance(val, list):
                return [int(x) if isinstance(x, int) else (ord(str(x).upper()[0]) - 65) for x in val]
            if isinstance(val, int):
                return val
            s = str(val).strip().upper()
            if s and s[0] in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
                return ord(s[0]) - 65
            try:
                return int(val)
            except (TypeError, ValueError):
                return None

        def is_correct(question_info, selected):
            correct = question_info.get('correct_answer')
            if correct is None:
                return False
            if isinstance(correct, list):
                return sorted(selected if isinstance(selected, list) else [selected]) == sorted(correct)
            return (selected if isinstance(selected, int) else (selected[0] if selected else None)) == correct

        created_attempts = {}
        participant_names = {}  # clicker_id or participant_id -> name for live app display
        answers_created = 0
        skipped_no_participant = 0
        skipped_no_question = 0
        skipped_already_answered = 0
        answers_updated = 0  # when revisable=True, overwrite existing answer
        for item in responses_data:
            participant = get_or_create_participant_for_clicker(
                participant_id=item.get('participant_id'),
                clicker_id=item.get('clicker_id')
            )
            if not participant:
                skipped_no_participant += 1
                continue
            cid = item.get('clicker_id')
            if cid is not None:
                participant_names[str(cid)] = participant.name
            participant_names[str(participant.id)] = participant.name
            question_id = item.get('question_id')
            if question_id is None:
                continue
            question_id = int(question_id)
            if question_id not in snapshot_questions:
                skipped_no_question += 1
                continue
            qinfo = snapshot_questions[question_id]
            selected = normalize_selected(item.get('selected_answer'))
            if selected is None:
                continue

            attempt_defaults = {'total_questions': len(snapshot_questions)}
            if exam_started_at_parsed is not None:
                attempt_defaults['started_at'] = exam_started_at_parsed
            attempt, attempt_created = ExamAttempt.objects.get_or_create(
                exam=exam,
                participant=participant,
                defaults=attempt_defaults
            )
            # Django ignores defaults for auto_now_add=True, so set started_at after create when client sent exam_started_at
            if attempt_created and exam_started_at_parsed is not None:
                attempt.started_at = exam_started_at_parsed
                attempt.save(update_fields=['started_at'])
                logger.info(
                    '[sync_live_results] TIME_TAKEN DEBUG: new attempt started_at set from client: %s',
                    exam_started_at_parsed
                )
            created_attempts[participant.id] = attempt

            correct = is_correct(qinfo, selected)
            pos = float(qinfo.get('positive_marks', 1))
            neg = float(qinfo.get('negative_marks', 0))
            time_taken = 0
            if item.get('answered_at') and attempt.started_at:
                try:
                    raw = item['answered_at']
                    if isinstance(raw, str):
                        from datetime import datetime
                        answered = datetime.fromisoformat(raw.replace('Z', '+00:00'))
                        if timezone.is_naive(answered):
                            answered = timezone.make_aware(answered)
                    else:
                        answered = raw
                    # Per-question time: delta from previous answer (or exam start for first answer)
                    prev = (
                        Answer.objects.filter(attempt=attempt, answered_at__lt=answered)
                        .order_by('-answered_at')
                        .values_list('answered_at', flat=True)
                        .first()
                    )
                    base = prev if prev else attempt.started_at
                    time_taken = max(0, int((answered - base).total_seconds()))
                    logger.info(
                        '[sync_live_results] TIME_TAKEN DEBUG: question_id=%s participant_id=%s '
                        'base=%s answered_at=%s -> time_taken(sec)=%s',
                        question_id, participant.id, base, raw, time_taken
                    )
                except Exception as e:
                    logger.warning('[sync_live_results] TIME_TAKEN DEBUG: failed to compute time_taken: %s', e)
            else:
                if item.get('answered_at') or attempt.started_at:
                    logger.info(
                        '[sync_live_results] TIME_TAKEN DEBUG: question_id=%s time_taken=0 (no answered_at or no attempt.started_at) '
                        'answered_at=%s attempt.started_at=%s',
                        question_id, item.get('answered_at'), getattr(attempt, 'started_at', None)
                    )

            existing = Answer.objects.filter(attempt=attempt, question_id=question_id).first()
            if existing:
                if exam.revisable:
                    existing.selected_answer = selected if isinstance(selected, list) else [selected]
                    existing.is_correct = correct
                    existing.time_taken = time_taken
                    existing.save()
                    answers_updated += 1
                else:
                    skipped_already_answered += 1
                continue

            Answer.objects.create(
                attempt=attempt,
                question_id=question_id,
                selected_answer=selected if isinstance(selected, list) else [selected],
                is_correct=correct,
                time_taken=time_taken
            )
            answers_created += 1

        logger.info(
            '[sync_live_results] Exam id=%s: answers_created=%d, answers_updated=%d, attempts_updated=%d, '
            'skipped_no_participant=%d, skipped_no_question=%d, skipped_already_answered=%d',
            pk, answers_created, answers_updated, len(created_attempts),
            skipped_no_participant, skipped_no_question, skipped_already_answered
        )

        # Recalculate attempt totals and submitted_at
        for attempt in created_attempts.values():
            answers = Answer.objects.filter(attempt=attempt)
            total = attempt.total_questions or 0
            correct_count = answers.filter(is_correct=True).count()
            wrong_count = answers.filter(is_correct=False).count()
            unattempted = max(0, total - answers.count())
            total_marks = sum(float(snapshot_questions.get(a.question_id, {}).get('positive_marks', 1)) for a in answers.filter(is_correct=True))
            total_marks -= sum(float(snapshot_questions.get(a.question_id, {}).get('negative_marks', 0)) for a in answers.filter(is_correct=False))
            attempt.correct_answers = correct_count
            attempt.wrong_answers = wrong_count
            attempt.unattempted = unattempted
            attempt.score = total_marks
            attempt.submitted_at = timezone.now()
            if answers.exists():
                last_ans = answers.order_by('-answered_at').first()
                if last_ans and attempt.started_at:
                    attempt.time_taken = max(0, int((last_ans.answered_at - attempt.started_at).total_seconds()))
            attempt.save()

        # Assign all listed participants to the exam and mark present those who responded
        for pid in attendance_ids:
            try:
                p = Participant.objects.get(id=pid)
                ExamParticipant.objects.get_or_create(exam=exam, participant=p)
                if pid in created_attempts:
                    continue
                if ExamAttempt.objects.filter(exam=exam, participant=p).exists():
                    continue
            except Participant.DoesNotExist:
                pass

        logger.info(
            '[sync_live_results] Exam id=%s: done. synced=%d, attempts_updated=%d',
            pk, answers_created, len(created_attempts)
        )
        return Response({
            'synced': answers_created,
            'answers_updated': answers_updated,
            'attempts_updated': len(created_attempts),
            'received': len(responses_data),
            'skipped_no_participant': skipped_no_participant,
            'skipped_no_question': skipped_no_question,
            'skipped_already_answered': skipped_already_answered,
            'participant_names': participant_names,
        })

    @action(detail=True, methods=['get'], url_path='attendance')
    def attendance(self, request, pk=None):
        """
        GET /api/exams/{id}/attendance/
        Returns attendance for the exam: list of participants with present/absent.
        Present = has an ExamAttempt for this exam (taken attendance or submitted answers).
        """
        exam = self.get_object()
        assigned = ExamParticipant.objects.filter(exam=exam).select_related('participant')
        present_ids = set(
            ExamAttempt.objects.filter(exam=exam).values_list('participant_id', flat=True)
        )
        participants = []
        for ep in assigned:
            p = ep.participant
            extra = p.extra or {}
            participants.append({
                'id': p.id,
                'name': p.name,
                'email': p.email,
                'clicker_id': p.clicker_id,
                'parent_email_id': extra.get('parent_email_id', '') if isinstance(extra, dict) else '',
                'parent_whatsapp': (
                    extra.get('parent_whatsapp')
                    or extra.get('parent_phone')
                    or extra.get('parent_mobile')
                    or ''
                ) if isinstance(extra, dict) else '',
                'present': p.id in present_ids,
            })
        present_count = len(present_ids)
        total_count = len(participants)
        return Response({
            'exam_id': exam.id,
            'exam_title': exam.title,
            'participants': participants,
            'present_count': present_count,
            'total_count': total_count,
        })

    @action(detail=True, methods=['get'], url_path='attendance/export')
    def attendance_export(self, request, pk=None):
        """
        GET /api/exams/{id}/attendance/export/?format=excel|pdf
        Downloads Attendance report as Excel or PDF.
        """
        exam = self.get_object()
        # IMPORTANT: DRF uses `?format=` for content negotiation; using it breaks this endpoint.
        # Use `file_format` instead (keep `format` as fallback for older clients).
        raw = (request.GET.get('file_format') or request.GET.get('format') or 'excel').strip().lower()
        format_type = 'pdf' if raw in ('pdf',) else 'excel'

        assigned = ExamParticipant.objects.filter(exam=exam).select_related('participant')
        present_ids = set(ExamAttempt.objects.filter(exam=exam).values_list('participant_id', flat=True))

        rows = []
        for ep in assigned:
            p = ep.participant
            extra = p.extra or {}
            status_label = 'Present' if p.id in present_ids else 'Absent'
            rows.append({
                'Name': p.name,
                'Keypad ID': p.clicker_id,
                'Email': p.email or '',
                'Parent Email ID': extra.get('parent_email_id', '') if isinstance(extra, dict) else '',
                'Parent WhatsApp': (
                    extra.get('parent_whatsapp')
                    or extra.get('parent_phone')
                    or extra.get('parent_mobile')
                    or ''
                ) if isinstance(extra, dict) else '',
                'Status': status_label,
            })

        filename_base = f'attendance-{exam.id}'
        if format_type == 'excel':
            output = BytesIO()
            df = pd.DataFrame(rows, columns=['Name', 'Keypad ID', 'Email', 'Parent Email ID', 'Parent WhatsApp', 'Status'])
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Attendance', index=False)
            output.seek(0)
            response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
            return response

        # PDF
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = [
            Paragraph(f'Attendance Report - {exam.title}', styles['Title']),
            Spacer(1, 12),
            Paragraph(f'Generated at: {timezone.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']),
            Spacer(1, 12),
        ]

        table_data = [['Name', 'Keypad ID', 'Email', 'Parent Email ID', 'Parent WhatsApp', 'Status']] + [
            [r['Name'], r['Keypad ID'], r['Email'], r['Parent Email ID'], r['Parent WhatsApp'], r['Status']] for r in rows
        ]
        table = Table(table_data, repeatRows=1, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(table)
        doc.build(elements)
        output.seek(0)

        response = HttpResponse(output.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
        return response

    @action(detail=True, methods=['post'], url_path='attendance/send-parent-emails')
    def send_parent_emails(self, request, pk=None):
        """
        POST /api/exams/{id}/attendance/send-parent-emails/
        Body:
          {
            "scope": "present" | "absent" | "all",
            "subject": "...",
            "body": ".... {{exam_title}} {{student_name}} {{clicker_id}} {{status}} ..."
          }
        Sends emails to Participant.extra.parent_email_id for assigned participants.
        """
        exam = self.get_object()
        scope = (request.data.get('scope') or 'absent').strip().lower()
        participant_ids = request.data.get('participant_ids', None)
        subject = (request.data.get('subject') or '').strip()
        body = (request.data.get('body') or '').strip()
        if scope not in ('present', 'absent', 'all'):
            return Response({'error': 'Invalid scope'}, status=status.HTTP_400_BAD_REQUEST)
        if participant_ids is not None and not isinstance(participant_ids, list):
            return Response({'error': 'participant_ids must be a list'}, status=status.HTTP_400_BAD_REQUEST)
        if not subject:
            return Response({'error': 'Subject is required'}, status=status.HTTP_400_BAD_REQUEST)
        if not body:
            return Response({'error': 'Body is required'}, status=status.HTTP_400_BAD_REQUEST)

        assigned = ExamParticipant.objects.filter(exam=exam).select_related('participant')
        if participant_ids:
            try:
                ids = [int(x) for x in participant_ids]
            except Exception:
                return Response({'error': 'participant_ids must be integers'}, status=status.HTTP_400_BAD_REQUEST)
            assigned = assigned.filter(participant_id__in=ids)
        present_ids = set(ExamAttempt.objects.filter(exam=exam).values_list('participant_id', flat=True))

        sent = 0
        skipped = 0
        errors = []

        # Reuse a single SMTP connection for all recipients.
        # This avoids long repeated reconnects and helps prevent gunicorn worker timeouts.
        connection = get_connection(fail_silently=False)
        try:
            connection.open()
        except Exception as e:
            # Network/SMTP connectivity failure (e.g., no internet, firewall blocking).
            # Return a normal JSON response so the UI can show the exact error.
            return Response({'sent': 0, 'skipped': 0, 'errors': [str(e)]}, status=status.HTTP_200_OK)

        def _render(template: str, ctx: dict) -> str:
            out = template
            for k, v in ctx.items():
                out = out.replace('{{' + k + '}}', str(v))
            return out

        try:
            for ep in assigned:
                p = ep.participant
                is_present = p.id in present_ids
                if scope == 'present' and not is_present:
                    continue
                if scope == 'absent' and is_present:
                    continue

                extra = p.extra or {}
                parent_email = extra.get('parent_email_id') if isinstance(extra, dict) else None
                parent_email = (parent_email or '').strip()
                if not parent_email:
                    skipped += 1
                    continue

                ctx = {
                    'exam_title': exam.title,
                    'student_name': p.name,
                    'clicker_id': p.clicker_id,
                    'status': 'Present' if is_present else 'Absent',
                }
                msg = EmailMessage(
                    subject=_render(subject, ctx),
                    body=_render(body, ctx),
                    to=[parent_email],
                    connection=connection,
                )
                try:
                    msg.send(fail_silently=False)
                    sent += 1
                except Exception as e:
                    errors.append(f'Participant {p.id}: {str(e)}')
        finally:
            try:
                connection.close()
            except Exception:
                pass

        return Response({'sent': sent, 'skipped': skipped, 'errors': errors})

    @action(detail=True, methods=['post'], url_path='attendance/send-parent-whatsapp')
    def send_parent_whatsapp(self, request, pk=None):
        """
        POST /api/exams/{id}/attendance/send-parent-whatsapp/
        Body:
          {
            "scope": "present" | "absent" | "all",
            "message": ".... {{exam_title}} {{student_name}} {{clicker_id}} {{status}} ...",
            "participant_ids": [1, 2, 3]  // optional
          }
        Returns generated WhatsApp links (wa.me) for selected participants.
        """
        exam = self.get_object()
        scope = (request.data.get('scope') or 'absent').strip().lower()
        participant_ids = request.data.get('participant_ids', None)
        message = (request.data.get('message') or '').strip()
        if scope not in ('present', 'absent', 'all'):
            return Response({'error': 'Invalid scope'}, status=status.HTTP_400_BAD_REQUEST)
        if participant_ids is not None and not isinstance(participant_ids, list):
            return Response({'error': 'participant_ids must be a list'}, status=status.HTTP_400_BAD_REQUEST)
        if not message:
            return Response({'error': 'Message is required'}, status=status.HTTP_400_BAD_REQUEST)

        assigned = ExamParticipant.objects.filter(exam=exam).select_related('participant')
        if participant_ids:
            try:
                ids = [int(x) for x in participant_ids]
            except Exception:
                return Response({'error': 'participant_ids must be integers'}, status=status.HTTP_400_BAD_REQUEST)
            assigned = assigned.filter(participant_id__in=ids)
        present_ids = set(ExamAttempt.objects.filter(exam=exam).values_list('participant_id', flat=True))

        sent = 0
        skipped = 0
        errors = []
        links = []

        def _render(template: str, ctx: dict) -> str:
            out = template
            for k, v in ctx.items():
                out = out.replace('{{' + k + '}}', str(v))
            return out

        def _normalize_phone(raw: str) -> str:
            # WhatsApp wa.me expects digits only (country code included).
            return ''.join(ch for ch in str(raw or '') if ch.isdigit())

        for ep in assigned:
            p = ep.participant
            is_present = p.id in present_ids
            if scope == 'present' and not is_present:
                continue
            if scope == 'absent' and is_present:
                continue

            extra = p.extra or {}
            parent_phone = None
            if isinstance(extra, dict):
                parent_phone = (
                    extra.get('parent_whatsapp')
                    or extra.get('parent_phone')
                    or extra.get('parent_mobile')
                )
            phone = _normalize_phone(parent_phone)
            if not phone:
                skipped += 1
                continue

            ctx = {
                'exam_title': exam.title,
                'student_name': p.name,
                'clicker_id': p.clicker_id,
                'status': 'Present' if is_present else 'Absent',
            }
            try:
                text = _render(message, ctx)
                wa_link = f'https://wa.me/{phone}?text={quote(text)}'
                links.append({
                    'participant_id': p.id,
                    'student_name': p.name,
                    'phone': phone,
                    'link': wa_link,
                })
                sent += 1
            except Exception as e:
                errors.append(f'Participant {p.id}: {str(e)}')

        return Response({'sent': sent, 'skipped': skipped, 'errors': errors, 'links': links})


def _parse_iso_date(date_str):
    if not date_str or not isinstance(date_str, str):
        return None
    try:
        from datetime import datetime
        return datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
    except ValueError:
        return None


def _daily_att_extra_row(p):
    extra = p.extra or {}
    if not isinstance(extra, dict):
        extra = {}
    return {
        'parent_email_id': extra.get('parent_email_id', '') or '',
        'parent_whatsapp': (
            extra.get('parent_whatsapp')
            or extra.get('parent_phone')
            or extra.get('parent_mobile')
            or ''
        ),
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def daily_attendance_summary(request):
    """List day-wise attendance stats for the last N calendar days (sidebar)."""
    from datetime import timedelta
    from collections import defaultdict

    try:
        days = int(request.query_params.get('days') or 60)
    except (TypeError, ValueError):
        days = 60
    days = max(1, min(days, 400))

    user = request.user
    pqs = scope_participants_queryset(Participant.objects.all(), user).order_by(
        Length('clicker_id'), 'clicker_id'
    )
    total = pqs.count()
    pid_list = list(pqs.values_list('id', flat=True))
    today = timezone.localdate()
    start = today - timedelta(days=days - 1)

    counts_by_date = defaultdict(lambda: {'present': 0, 'absent': 0})
    if pid_list:
        for row in DailyAttendance.objects.filter(
            participant_id__in=pid_list, date__gte=start, date__lte=today
        ).values('date', 'present'):
            d = row['date']
            if row['present']:
                counts_by_date[d]['present'] += 1
            else:
                counts_by_date[d]['absent'] += 1

    out = []
    for i in range(days):
        d = today - timedelta(days=i)
        c = counts_by_date[d]
        pr, ab = c['present'], c['absent']
        marked = pr + ab
        out.append({
            'date': d.isoformat(),
            'present_count': pr,
            'absent_count': ab,
            'unmarked_count': max(0, total - marked),
            'total_count': total,
        })
    return Response(out)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def daily_attendance_day(request):
    """Roster for one day: all scoped participants with daily mark (if any)."""
    d = _parse_iso_date(request.query_params.get('date'))
    if not d:
        return Response({'error': 'Invalid or missing date (use YYYY-MM-DD).'}, status=status.HTTP_400_BAD_REQUEST)

    user = request.user
    pqs = scope_participants_queryset(Participant.objects.all(), user).order_by(
        Length('clicker_id'), 'clicker_id'
    )
    pids = list(pqs.values_list('id', flat=True))
    records = {}
    if pids:
        records = {
            r.participant_id: r
            for r in DailyAttendance.objects.filter(date=d, participant_id__in=pids)
        }

    participants = []
    present_n = absent_n = unmarked_n = 0
    for p in pqs:
        rec = records.get(p.id)
        marked = rec is not None
        present = bool(rec.present) if rec else False
        if not marked:
            unmarked_n += 1
        elif present:
            present_n += 1
        else:
            absent_n += 1
        ex = _daily_att_extra_row(p)
        participants.append({
            'id': p.id,
            'name': p.name,
            'email': p.email or '',
            'clicker_id': p.clicker_id,
            'parent_email_id': ex['parent_email_id'],
            'parent_whatsapp': ex['parent_whatsapp'],
            'present': present,
            'marked': marked,
        })

    return Response({
        'date': d.isoformat(),
        'participants': participants,
        'present_count': present_n,
        'absent_count': absent_n,
        'unmarked_count': unmarked_n,
        'total_count': len(participants),
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def daily_attendance_save(request):
    """Save attendance for a day.

    Body: { \"date\": \"YYYY-MM-DD\", \"entries\": [{\"participant_id\": 1, \"status\": \"present\"|\"absent\"|\"unmarked\"}, ...] }
    Legacy: { \"participant_id\", \"present\": true/false } treats false as absent and true as present.
    """
    d = _parse_iso_date((request.data or {}).get('date'))
    if not d:
        return Response({'error': 'Invalid or missing date (use YYYY-MM-DD).'}, status=status.HTTP_400_BAD_REQUEST)
    entries = (request.data or {}).get('entries')
    if not isinstance(entries, list) or not entries:
        return Response({'error': 'entries must be a non-empty list.'}, status=status.HTTP_400_BAD_REQUEST)

    user = request.user
    allowed_ids = set(
        scope_participants_queryset(Participant.objects.all(), user).values_list('id', flat=True)
    )
    saved = 0
    errors = []
    for item in entries:
        if not isinstance(item, dict):
            errors.append('Invalid entry')
            continue
        try:
            pid = int(item.get('participant_id'))
        except (TypeError, ValueError):
            errors.append('Invalid participant_id')
            continue
        if pid not in allowed_ids:
            errors.append(f'Participant {pid} not in scope')
            continue

        status_val = item.get('status')
        if status_val is None and 'present' in item:
            status_val = 'present' if bool(item.get('present')) else 'absent'
        status_val = (str(status_val or '')).strip().lower()
        if status_val == 'unmarked':
            DailyAttendance.objects.filter(participant_id=pid, date=d).delete()
            saved += 1
            continue
        if status_val in ('present', 'absent'):
            DailyAttendance.objects.update_or_create(
                participant_id=pid,
                date=d,
                defaults={'present': status_val == 'present', 'recorded_by': user},
            )
            saved += 1
        else:
            errors.append(f'Participant {pid}: invalid status')
    return Response({'saved': saved, 'errors': errors})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def daily_attendance_export(request):
    """GET /api/attendance/day/export/?date=YYYY-MM-DD&file_format=excel|pdf"""
    d = _parse_iso_date(request.query_params.get('date'))
    if not d:
        return Response({'error': 'Invalid or missing date (use YYYY-MM-DD).'}, status=status.HTTP_400_BAD_REQUEST)

    user = request.user
    pqs = scope_participants_queryset(Participant.objects.all(), user).order_by(
        Length('clicker_id'), 'clicker_id'
    )
    pids = list(pqs.values_list('id', flat=True))
    records = {}
    if pids:
        records = {
            r.participant_id: r
            for r in DailyAttendance.objects.filter(date=d, participant_id__in=pids)
        }

    rows = []
    for p in pqs:
        rec = records.get(p.id)
        ex = _daily_att_extra_row(p)
        if rec is None:
            status_label = 'Not recorded'
        else:
            status_label = 'Present' if rec.present else 'Absent'
        rows.append({
            'Name': p.name,
            'Keypad ID': p.clicker_id,
            'Email': p.email or '',
            'Parent Email ID': ex['parent_email_id'],
            'Parent WhatsApp': ex['parent_whatsapp'],
            'Status': status_label,
        })

    raw = (request.GET.get('file_format') or request.GET.get('format') or 'excel').strip().lower()
    format_type = 'pdf' if raw in ('pdf',) else 'excel'
    filename_base = f'attendance-daily-{d.isoformat()}'

    if format_type == 'excel':
        output = BytesIO()
        df = pd.DataFrame(rows, columns=['Name', 'Keypad ID', 'Email', 'Parent Email ID', 'Parent WhatsApp', 'Status'])
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Attendance', index=False)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
        return response

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    styles = getSampleStyleSheet()
    title = f'Daily Attendance — {d.strftime("%d/%m/%Y")}'
    elements = [
        Paragraph(title, styles['Title']),
        Spacer(1, 12),
        Paragraph(f'Generated at: {timezone.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']),
        Spacer(1, 12),
    ]
    table_data = [['Name', 'Keypad ID', 'Email', 'Parent Email ID', 'Parent WhatsApp', 'Status']] + [
        [r['Name'], r['Keypad ID'], r['Email'], r['Parent Email ID'], r['Parent WhatsApp'], r['Status']] for r in rows
    ]
    table = Table(table_data, repeatRows=1, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def daily_attendance_send_parent_emails(request):
    """Same body as exam attendance emails; uses daily marks for `date`."""
    d = _parse_iso_date((request.data or {}).get('date'))
    if not d:
        return Response({'error': 'Invalid or missing date (use YYYY-MM-DD).'}, status=status.HTTP_400_BAD_REQUEST)

    scope = ((request.data or {}).get('scope') or 'absent').strip().lower()
    participant_ids = (request.data or {}).get('participant_ids', None)
    subject = ((request.data or {}).get('subject') or '').strip()
    body = ((request.data or {}).get('body') or '').strip()
    if scope not in ('present', 'absent', 'all', 'unmarked'):
        return Response({'error': 'Invalid scope'}, status=status.HTTP_400_BAD_REQUEST)
    if participant_ids is not None and not isinstance(participant_ids, list):
        return Response({'error': 'participant_ids must be a list'}, status=status.HTTP_400_BAD_REQUEST)
    if not subject:
        return Response({'error': 'Subject is required'}, status=status.HTTP_400_BAD_REQUEST)
    if not body:
        return Response({'error': 'Body is required'}, status=status.HTTP_400_BAD_REQUEST)

    user = request.user
    pqs = scope_participants_queryset(Participant.objects.all(), user).order_by(
        Length('clicker_id'), 'clicker_id'
    )
    if participant_ids:
        try:
            ids = [int(x) for x in participant_ids]
        except Exception:
            return Response({'error': 'participant_ids must be integers'}, status=status.HTTP_400_BAD_REQUEST)
        pqs = pqs.filter(id__in=ids)

    pids = list(pqs.values_list('id', flat=True))
    present_ids = set()
    marked_ids = set()
    if pids:
        for r in DailyAttendance.objects.filter(date=d, participant_id__in=pids).values(
            'participant_id', 'present'
        ):
            marked_ids.add(r['participant_id'])
            if r['present']:
                present_ids.add(r['participant_id'])

    attendance_date_label = d.strftime('%d/%m/%Y')
    sent = 0
    skipped = 0
    errors = []
    connection = get_connection(fail_silently=False)
    try:
        connection.open()
    except Exception as e:
        return Response({'sent': 0, 'skipped': 0, 'errors': [str(e)]}, status=status.HTTP_200_OK)

    def _render(template: str, ctx: dict) -> str:
        out = template
        for k, v in ctx.items():
            out = out.replace('{{' + k + '}}', str(v))
        return out

    try:
        for p in pqs:
            is_marked = p.id in marked_ids
            is_present = p.id in present_ids
            if scope == 'unmarked' and is_marked:
                continue
            if scope == 'present' and (not is_marked or not is_present):
                continue
            if scope == 'absent' and (not is_marked or is_present):
                continue

            extra = _daily_att_extra_row(p)
            parent_email = (extra.get('parent_email_id') or '').strip()
            if not parent_email:
                skipped += 1
                continue
            status_human = 'Present' if is_present else ('Absent' if is_marked else 'Not recorded')
            ctx = {
                'exam_title': '',
                'attendance_date': attendance_date_label,
                'student_name': p.name,
                'clicker_id': p.clicker_id,
                'status': status_human,
            }
            msg = EmailMessage(
                subject=_render(subject, ctx),
                body=_render(body, ctx),
                to=[parent_email],
                connection=connection,
            )
            try:
                msg.send(fail_silently=False)
                sent += 1
            except Exception as e:
                errors.append(f'Participant {p.id}: {str(e)}')
    finally:
        try:
            connection.close()
        except Exception:
            pass

    return Response({'sent': sent, 'skipped': skipped, 'errors': errors})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def daily_attendance_send_parent_whatsapp(request):
    d = _parse_iso_date((request.data or {}).get('date'))
    if not d:
        return Response({'error': 'Invalid or missing date (use YYYY-MM-DD).'}, status=status.HTTP_400_BAD_REQUEST)

    scope = ((request.data or {}).get('scope') or 'absent').strip().lower()
    participant_ids = (request.data or {}).get('participant_ids', None)
    message = ((request.data or {}).get('message') or '').strip()
    if scope not in ('present', 'absent', 'all', 'unmarked'):
        return Response({'error': 'Invalid scope'}, status=status.HTTP_400_BAD_REQUEST)
    if participant_ids is not None and not isinstance(participant_ids, list):
        return Response({'error': 'participant_ids must be a list'}, status=status.HTTP_400_BAD_REQUEST)
    if not message:
        return Response({'error': 'Message is required'}, status=status.HTTP_400_BAD_REQUEST)

    user = request.user
    pqs = scope_participants_queryset(Participant.objects.all(), user).order_by(
        Length('clicker_id'), 'clicker_id'
    )
    if participant_ids:
        try:
            ids = [int(x) for x in participant_ids]
        except Exception:
            return Response({'error': 'participant_ids must be integers'}, status=status.HTTP_400_BAD_REQUEST)
        pqs = pqs.filter(id__in=ids)

    pids = list(pqs.values_list('id', flat=True))
    present_ids = set()
    marked_ids = set()
    if pids:
        for r in DailyAttendance.objects.filter(date=d, participant_id__in=pids).values('participant_id', 'present'):
            marked_ids.add(r['participant_id'])
            if r['present']:
                present_ids.add(r['participant_id'])

    attendance_date_label = d.strftime('%d/%m/%Y')
    sent = 0
    skipped = 0
    errors = []
    links = []

    def _render(template: str, ctx: dict) -> str:
        out = template
        for k, v in ctx.items():
            out = out.replace('{{' + k + '}}', str(v))
        return out

    def _normalize_phone(raw: str) -> str:
        return ''.join(ch for ch in str(raw or '') if ch.isdigit())

    for p in pqs:
        is_marked = p.id in marked_ids
        is_present = p.id in present_ids
        if scope == 'unmarked' and is_marked:
            continue
        if scope == 'present' and (not is_marked or not is_present):
            continue
        if scope == 'absent' and (not is_marked or is_present):
            continue

        extra = _daily_att_extra_row(p)
        phone = _normalize_phone(extra.get('parent_whatsapp'))
        if not phone:
            skipped += 1
            continue
        status_human = 'Present' if is_present else ('Absent' if is_marked else 'Not recorded')
        ctx = {
            'exam_title': '',
            'attendance_date': attendance_date_label,
            'student_name': p.name,
            'clicker_id': p.clicker_id,
            'status': status_human,
        }
        try:
            text = _render(message, ctx)
            wa_link = f'https://wa.me/{phone}?text={quote(text)}'
            links.append({
                'participant_id': p.id,
                'student_name': p.name,
                'phone': phone,
                'link': wa_link,
            })
            sent += 1
        except Exception as e:
            errors.append(f'Participant {p.id}: {str(e)}')

    return Response({'sent': sent, 'skipped': skipped, 'errors': errors, 'links': links})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_question_media(request):
    """Upload image or video for question rich text (stored under MEDIA_ROOT). Returns JSON { url }."""
    from django.core.files.storage import default_storage

    uploaded = request.FILES.get('file')
    if not uploaded:
        return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

    orig = (uploaded.name or 'file').strip()
    ext = os.path.splitext(orig)[1].lower()
    if not ext:
        return Response({'detail': 'File must have an extension.'}, status=status.HTTP_400_BAD_REQUEST)

    allowed_img = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
    allowed_vid = {'.mp4', '.webm', '.ogg', '.mov', '.m4v'}
    is_img = ext in allowed_img
    is_vid = ext in allowed_vid
    if not is_img and not is_vid:
        return Response(
            {
                'detail': f'Unsupported file type. Images: {sorted(allowed_img)}. Video: {sorted(allowed_vid)}.',
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    max_image = 15 * 1024 * 1024
    max_video = 120 * 1024 * 1024
    max_bytes = max_image if is_img else max_video
    if uploaded.size > max_bytes:
        return Response({'detail': 'File too large.'}, status=status.HTTP_400_BAD_REQUEST)

    sub = 'images' if is_img else 'videos'
    key = f'question_media/{sub}/{uuid.uuid4().hex}{ext}'
    saved_name = default_storage.save(key, uploaded)
    relative = default_storage.url(saved_name)
    full_url = request.build_absolute_uri(relative)
    return Response({'url': full_url})


# Question Views
class QuestionViewSet(viewsets.ModelViewSet):
    serializer_class = QuestionSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        """Set created_by when a question is created manually."""
        serializer.save(created_by=self.request.user)

    def get_queryset(self):
        """Scope questions by role.

        - Super Admin: all questions
        - School Admin: questions created by users in their school
        - Teacher: only questions they created
        """
        qs = Question.objects.all()
        user = self.request.user
        role = get_user_role(user)
        params = self.request.query_params

        def _parse_int(raw):
            try:
                return int(raw)
            except (TypeError, ValueError):
                return None

        school_id_param = _parse_int(params.get('school_id'))
        teacher_id_param = _parse_int(params.get('teacher_id') or params.get('created_by'))

        # RBAC scope first
        if role == ROLE_SUPER_ADMIN:
            # optional filter for super admin
            if school_id_param is not None:
                qs = qs.filter(created_by__profile__school_id=school_id_param)
            if teacher_id_param is not None:
                qs = qs.filter(created_by_id=teacher_id_param)
            return qs

        if role == ROLE_SCHOOL_ADMIN:
            user_school_id = get_user_school_id(user)
            if not user_school_id:
                return qs.none()
            qs = qs.filter(created_by__profile__school_id=user_school_id)
            # allow further narrowing
            if school_id_param is not None and school_id_param != user_school_id:
                return qs.none()
            if teacher_id_param is not None:
                qs = qs.filter(created_by_id=teacher_id_param)
            return qs

        if role == ROLE_TEACHER:
            # teacher can only see their own questions; if they ask for another teacher, return none
            if teacher_id_param is not None and teacher_id_param != user.id:
                return qs.none()
            return qs.filter(created_by=user)

        return qs.none()

    def _normalize_topic(self, raw: str) -> str:
        """Ensure topic is a short label, not a full prompt (e.g. from mis-sent request)."""
        if not raw:
            return raw
        topic = raw.strip()
        # Reject prompt-like or overly long input: use only first line or first 200 chars
        if len(topic) > 200 or topic.lower().startswith('you are ') or 'generate' in topic.lower()[:80]:
            first_line = topic.split('\n')[0].strip()
            if len(first_line) > 200:
                topic = first_line[:200].strip()
            else:
                topic = first_line
            # If it still looks like a prompt, use a safe fallback
            if 'generate' in topic.lower() or topic.lower().startswith('you are'):
                topic = 'General'
        return topic[:200]

    @action(detail=False, methods=['post'])
    def generate(self, request):
        """Generate questions using AI"""
        import re
        raw_topic = request.data.get('topic', '').strip()
        topic = self._normalize_topic(raw_topic)
        count = int(request.data.get('count', 5))
        difficulty = request.data.get('difficulty', 'medium')
        qtype = request.data.get('type', 'mcq')
        # Optional: number of options per question (MCQ/multiple_select). Default 4; can parse from topic e.g. "with 10 options"
        num_options = request.data.get('num_options')
        if num_options is not None:
            num_options = int(num_options)
            if num_options < 2 or num_options > 15:
                num_options = 4
        else:
            # Try to parse from topic: "with 10 options", "10 options", "class 6 math with 10 options"
            match = re.search(r'(?:with\s+)?(\d+)\s*options?', raw_topic, re.IGNORECASE)
            if match:
                n = int(match.group(1))
                if 2 <= n <= 15:
                    num_options = n
            if num_options is None:
                num_options = 4
        # If we parsed num_options from topic, strip that phrase so the topic is cleaner for the AI
        if request.data.get('num_options') is None and re.search(r'(?:with\s+)?\d+\s*options?', raw_topic, re.IGNORECASE):
            topic = re.sub(r'\s*with\s+\d+\s*options?\s*', ' ', topic, flags=re.IGNORECASE)
            topic = re.sub(r'\s*\d+\s*options?\s*', ' ', topic, flags=re.IGNORECASE)
            topic = re.sub(r'\s+', ' ', topic).strip() or topic
        
        # Validate inputs
        if not topic:
            return Response(
                {'error': 'Topic is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if count < 1 or count > 20:
            return Response(
                {'error': 'Count must be between 1 and 20'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if difficulty not in ['easy', 'medium', 'hard']:
            return Response(
                {'error': 'Difficulty must be easy, medium, or hard'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if qtype not in ['mcq', 'true_false', 'multiple_select']:
            return Response(
                {'error': 'Type must be mcq, true_false, or multiple_select'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Option label format: alpha (A,B,C) or numeric (1,2,3). Support both snake_case and camelCase.
        option_display = request.data.get('option_display') or request.data.get('optionDisplay') or 'alpha'
        if isinstance(option_display, str):
            option_display = option_display.strip().lower()
        if option_display not in ('alpha', 'numeric'):
            option_display = 'alpha'

        try:
            # Try AI generation: Groq first, then Gemini (free), then OpenAI
            ai_questions = []
            provider_used = None
            try:
                from api.services.ai_generator_groq import GroqQuestionGenerator
                generator = GroqQuestionGenerator()
                ai_questions = generator.generate_questions_safe(topic, count, difficulty, qtype, num_options=num_options)
                if ai_questions:
                    provider_used = 'Groq'
            except (ValueError, ImportError) as e:
                logger.info("AI generate: Groq not used (%s)", e)
                pass

            if not ai_questions:
                try:
                    from api.services.ai_generator_gemini import GeminiQuestionGenerator
                    generator = GeminiQuestionGenerator()
                    ai_questions = generator.generate_questions_safe(topic, count, difficulty, qtype, num_options=num_options)
                    if ai_questions:
                        provider_used = 'Gemini'
                except (ValueError, ImportError) as e:
                    logger.info("AI generate: Gemini not used (%s)", e)
                    pass

            if not ai_questions:
                try:
                    from api.services.ai_generator import AIQuestionGenerator
                    generator = AIQuestionGenerator()
                    ai_questions = generator.generate_questions_safe(topic, count, difficulty, qtype, num_options=num_options)
                    if ai_questions:
                        provider_used = 'OpenAI'
                except (ValueError, ImportError) as e:
                    logger.info("AI generate: OpenAI not used (%s)", e)
                    pass

            # If AI generation succeeded and returned questions
            if ai_questions:
                logger.info(
                    "AI generate: %s generated %d questions for topic=%s",
                    provider_used or "AI", len(ai_questions), topic[:50],
                )
                print(
                    f"[EasyTest AI] Generated {len(ai_questions)} questions using {provider_used or 'AI'} for topic: {topic[:60]}..."
                )
                generated_questions = []
                for q_data in ai_questions:
                    question = Question.objects.create(
                        text=q_data['text'],
                        type=qtype,
                        options=q_data['options'],
                        correct_answer=q_data['correct_answer'],
                        option_display=option_display,
                        difficulty=q_data.get('difficulty', difficulty),
                        tags=q_data.get('tags', [topic]),
                        marks=q_data.get('marks', 1.0),
                        created_by=request.user,
                    )
                    generated_questions.append(QuestionSerializer(question).data)
                
                return Response(generated_questions, status=status.HTTP_201_CREATED)
            else:
                # Fallback to mock if AI fails or no provider configured
                logger.warning(
                    "AI generate: no provider available or all returned empty; using sample questions for topic=%s",
                    topic[:50],
                )
                print(
                    "[EasyTest AI] No API key configured or AI returned no questions. "
                    "Using sample questions. For real AI: add GROQ_API_KEY or GEMINI_API_KEY to .env (see FREE_AI_SETUP.md)"
                )
                mock_data = self._generate_mock_questions(topic, count, difficulty, qtype)
                return Response(
                    {
                        'questions': mock_data,
                        'warning': 'AI generation unavailable. Sample questions generated. Add GROQ_API_KEY or GEMINI_API_KEY to .env for real AI (see FREE_AI_SETUP.md).'
                    },
                    status=status.HTTP_201_CREATED
                )
                
        except ValueError as e:
            # API key not set - use mock
            logger.warning("AI generate: ValueError (e.g. missing API key): %s", e)
            print(
                "[EasyTest AI] API key not set. Using sample questions. "
                "Add GROQ_API_KEY or GEMINI_API_KEY to backend .env (see FREE_AI_SETUP.md)"
            )
            mock_data = self._generate_mock_questions(topic, count, difficulty, qtype)
            return Response(
                {
                    'questions': mock_data,
                    'warning': 'API key not configured. Sample questions generated. Add GROQ_API_KEY or GEMINI_API_KEY to .env for real AI.'
                },
                status=status.HTTP_201_CREATED
            )
        except Exception as e:
            # Any other error - use mock
            error_msg = str(e)
            logger.error("AI generation error: %s", error_msg)
            print(f"[EasyTest AI] Error: {error_msg[:200]}. Using sample questions.")
            
            # Check for specific error types
            warning_msg = 'AI generation unavailable. Sample questions generated.'
            if 'quota' in error_msg.lower() or '429' in error_msg:
                warning_msg = 'API quota exceeded. Sample questions generated.'
            elif 'api key' in error_msg.lower() or '401' in error_msg:
                warning_msg = 'Invalid API key. Sample questions generated.'
            
            mock_data = self._generate_mock_questions(topic, count, difficulty, qtype)
            return Response(
                {
                    'questions': mock_data,
                    'warning': warning_msg
                },
                status=status.HTTP_201_CREATED
            )
    
    def _generate_mock_questions(self, topic: str, count: int, difficulty: str, qtype: str):
        """Fallback mock question generation"""
        # Keep topic short for display (avoid prompt-sized strings)
        safe_topic = (topic[:100] + '…') if len(topic) > 100 else topic
        generated_questions = []
        for i in range(count):
            question = Question.objects.create(
                text=f"Sample question about {safe_topic} - Question {i+1}",
                type=qtype,
                options=['Option A', 'Option B', 'Option C', 'Option D'] if qtype == 'mcq' else ['True', 'False'],
                correct_answer=0 if qtype == 'mcq' else [0],
                difficulty=difficulty,
                tags=[safe_topic],
                marks=1.0
            )
            generated_questions.append(QuestionSerializer(question).data)
        
        return generated_questions  # Return list directly

    @action(detail=False, methods=['post'], url_path='import')
    def import_questions(self, request):
        """Import questions from CSV or Excel. Required: text, options, correct_answer. Optional: type, difficulty, marks, tags."""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            elif file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file)
            else:
                return Response({'error': 'Unsupported file format. Use CSV or Excel.'}, status=status.HTTP_400_BAD_REQUEST)

            def get_col(df, *candidates):
                for c in candidates:
                    for col in df.columns:
                        if str(col).strip().lower() == c.lower():
                            return col
                return None

            text_col = get_col(df, 'text', 'question', 'question text', 'questions')
            options_col = get_col(df, 'options', 'option', 'choices')
            correct_col = get_col(df, 'correct_answer', 'correct answer', 'correct', 'answer', 'answer_index')
            type_col = get_col(df, 'type', 'question type', 'qtype')
            difficulty_col = get_col(df, 'difficulty')
            marks_col = get_col(df, 'marks', 'mark', 'points')
            tags_col = get_col(df, 'tags', 'tag')
            option_display_col = get_col(df, 'option_display', 'option display', 'options format', 'label format')

            if text_col is None or options_col is None or correct_col is None:
                return Response(
                    {'error': 'File must have columns: text (or question), options, correct_answer (or correct/answer).'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            imported = 0
            errors = []
            for idx, row in df.iterrows():
                try:
                    raw_text = row.get(text_col, '')
                    text = '' if pd.isna(raw_text) else str(raw_text).strip()
                    if not text:
                        errors.append(f"Row {idx + 2}: Question text is required.")
                        continue

                    raw_opts = row.get(options_col, '')
                    opts_str = '' if pd.isna(raw_opts) else str(raw_opts).strip()
                    if not opts_str:
                        errors.append(f"Row {idx + 2}: Options are required (use pipe | or semicolon ; to separate).")
                        continue
                    options = [o.strip() for o in opts_str.replace(';', '|').split('|') if o.strip()]
                    if len(options) < 2:
                        errors.append(f"Row {idx + 2}: At least 2 options required.")
                        continue

                    raw_correct = row.get(correct_col, '')
                    correct_str = '' if pd.isna(raw_correct) else str(raw_correct).strip()
                    if correct_str == '':
                        errors.append(f"Row {idx + 2}: correct_answer is required (0-based index or comma-separated for multiple select).")
                        continue
                    if ',' in correct_str:
                        correct_answer = [int(x.strip()) for x in correct_str.split(',') if x.strip() != '']
                        if not correct_answer or any(i < 0 or i >= len(options) for i in correct_answer):
                            errors.append(f"Row {idx + 2}: correct_answer indices must be 0 to {len(options) - 1}.")
                            continue
                    else:
                        try:
                            correct_answer = int(correct_str)
                        except ValueError:
                            errors.append(f"Row {idx + 2}: correct_answer must be a number or comma-separated numbers.")
                            continue
                        if correct_answer < 0 or correct_answer >= len(options):
                            errors.append(f"Row {idx + 2}: correct_answer must be 0 to {len(options) - 1}.")
                            continue

                    qtype = 'mcq'
                    if type_col:
                        raw_type = row.get(type_col, '')
                        t = '' if pd.isna(raw_type) else str(raw_type).strip().lower()
                        if t in ('mcq', 'true_false', 'multiple_select'):
                            qtype = t
                        elif t in ('tf', 'true false'):
                            qtype = 'true_false'
                        elif t in ('ms', 'multiple select', 'multi'):
                            qtype = 'multiple_select'

                    diff = 'medium'
                    if difficulty_col:
                        raw_diff = row.get(difficulty_col, '')
                        d = '' if pd.isna(raw_diff) else str(raw_diff).strip().lower()
                        if d in ('easy', 'medium', 'hard'):
                            diff = d

                    marks_val = 1.0
                    if marks_col:
                        raw_marks = row.get(marks_col, '')
                        if pd.notna(raw_marks):
                            try:
                                marks_val = float(raw_marks)
                            except (ValueError, TypeError):
                                pass

                    tags_list = []
                    if tags_col:
                        raw_tags = row.get(tags_col, '')
                        if pd.notna(raw_tags) and str(raw_tags).strip():
                            tags_list = [t.strip() for t in str(raw_tags).split(',') if t.strip()]

                    option_display = 'alpha'
                    if option_display_col:
                        raw_od = row.get(option_display_col, '')
                        od = '' if pd.isna(raw_od) else str(raw_od).strip().lower()
                        if od in ('numeric', 'number', 'num', '1'):
                            option_display = 'numeric'
                        elif od in ('alpha', 'letter', 'a', 'abc'):
                            option_display = 'alpha'

                    Question.objects.create(
                        text=text,
                        type=qtype,
                        options=options,
                        correct_answer=correct_answer,
                        option_display=option_display,
                        difficulty=diff,
                        tags=tags_list,
                        marks=marks_val
                    )
                    imported += 1
                except Exception as e:
                    errors.append(f"Row {idx + 2}: {str(e)}")

            return Response({'imported': imported, 'errors': errors})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# Participant Views
class ParticipantViewSet(viewsets.ModelViewSet):
    queryset = Participant.objects.all()
    serializer_class = ParticipantSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        exam_id = self.request.query_params.get('exam_id')
        school_id_param = self.request.query_params.get('school_id')
        teacher_id_param = (
            self.request.query_params.get('teacher_id')
            or self.request.query_params.get('created_by')
        )
        user = self.request.user
        base_order = [Length('clicker_id'), 'clicker_id']
        base_qs = scope_participants_queryset(Participant.objects.all(), user)
        role = get_user_role(user)
        if school_id_param not in (None, ''):
            try:
                sid = int(school_id_param)
            except (TypeError, ValueError):
                sid = None
            if sid is not None:
                if role == ROLE_SCHOOL_ADMIN:
                    user_school = get_user_school_id(user)
                    if user_school is not None and sid != user_school:
                        base_qs = base_qs.none()
                    else:
                        base_qs = base_qs.filter(school_id=sid)
                else:
                    base_qs = base_qs.filter(school_id=sid)
        if teacher_id_param not in (None, ''):
            try:
                tid = int(teacher_id_param)
            except (TypeError, ValueError):
                tid = None
            if tid is not None:
                base_qs = base_qs.filter(created_by_id=tid)
        if exam_id:
            exam_qs = scope_exams_queryset(Exam.objects.all(), user)
            try:
                exam = exam_qs.get(id=exam_id)
            except Exam.DoesNotExist:
                return Participant.objects.none()
            participant_ids = ExamParticipant.objects.filter(exam=exam).values_list('participant_id', flat=True)
            return base_qs.filter(id__in=participant_ids).order_by(*base_order)
        return base_qs.order_by(*base_order)

    def perform_create(self, serializer):
        """Set created_by and school on single participant create."""
        school_id = get_user_school_id(self.request.user)
        serializer.save(created_by=self.request.user, school_id=school_id)

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Create multiple participants in one request. Each item needs name and clicker_id; email optional."""
        from .serializers import ParticipantBulkCreateSerializer
        ser = ParticipantBulkCreateSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        RESERVED_KEYS = {'name', 'clicker_id', 'email'}
        created = []
        errors = []
        for i, item in enumerate(ser.validated_data['participants']):
            name = (item.get('name') or '').strip()
            clicker_id = (item.get('clicker_id') or '').strip()
            email = (item.get('email') or '').strip() or None
            extra = {k: (v if isinstance(v, str) else str(v)) for k, v in item.items()
                     if k not in RESERVED_KEYS and v is not None and str(v).strip() != ''}
            try:
                school_id = get_user_school_id(request.user)
                participant = Participant.objects.create(
                    name=name,
                    clicker_id=clicker_id,
                    email=email,
                    school_id=school_id,
                    extra=extra,
                    created_by=request.user,
                )
                created.append(ParticipantSerializer(participant).data)
            except Exception as e:
                errors.append(f"Row {i + 1}: {str(e)}")
        return Response(
            {'created': len(created), 'participants': created, 'errors': errors},
            status=status.HTTP_201_CREATED
        )

    @action(detail=False, methods=['post'], url_path='import')
    def import_participants(self, request):
        file = request.FILES.get('file')
        exam_id = request.data.get('exam_id')
        
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            elif file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file)
            else:
                return Response({'error': 'Unsupported file format'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Flexible column mapping: name and clicker_id required; email optional
            # Normalize optional column headers to canonical keys for extra (matches frontend form)
            def _normalize_col_key(col_name):
                if col_name is None or (isinstance(col_name, float) and pd.isna(col_name)):
                    return None
                s = str(col_name).strip().lower()
                COLUMN_TO_EXTRA_KEY = {
                    'roll no': 'roll_no', 'roll no.': 'roll_no', 'roll number': 'roll_no', 'roll number.': 'roll_no',
                    'admission no': 'admission_no', 'admission no.': 'admission_no', 'admission number': 'admission_no', 'admission number.': 'admission_no',
                    'class': 'class',
                    'subject': 'subject',
                    'section': 'section',
                    'team': 'team',
                    'group': 'group',
                    'house': 'house',
                    'gender': 'gender',
                    'city': 'city',
                    'uid': 'uid',
                    'employee code': 'employee_code', 'employee code.': 'employee_code',
                    'teacher name': 'teacher_name', 'teacher': 'teacher_name', 'teaccher name': 'teacher_name',
                    'email id': 'email_id', 'email': 'email_id', 'e-mail': 'email_id', 'email address': 'email_id',
                    'parent email id': 'parent_email_id', 'parent email': 'parent_email_id',
                    'parent e-mail': 'parent_email_id', 'parent e mail': 'parent_email_id',
                    'parent_email': 'parent_email_id', 'parent_email_id': 'parent_email_id',
                    'parents email': 'parent_email_id', 'parents email id': 'parent_email_id',
                    'father email': 'parent_email_id', 'mother email': 'parent_email_id',
                    'guardian email': 'parent_email_id',
                    'parent whatsapp': 'parent_whatsapp',
                    'parent whatsapp number': 'parent_whatsapp',
                    'parent whatsapp no': 'parent_whatsapp',
                    'parent phone': 'parent_whatsapp',
                    'parent mobile': 'parent_whatsapp',
                    'parent mobile number': 'parent_whatsapp',
                    'guardian phone': 'parent_whatsapp',
                    'guardian mobile': 'parent_whatsapp',
                    'whatsapp': 'parent_whatsapp',
                    'whatsapp number': 'parent_whatsapp',
                }
                return COLUMN_TO_EXTRA_KEY.get(s) or s.replace(' ', '_').replace('.', '').replace('-', '_') or None

            def get_col(df, *candidates):
                for c in candidates:
                    for col in df.columns:
                        if str(col).strip().lower() == c.lower():
                            return col
                return None
            name_col = get_col(df, 'name', 'names', 'participant', 'participant name')
            clicker_col = get_col(df, 'clicker_id', 'clicker id', 'clickerid', 'clicker', 'keypad id', 'keypadid', 'keypad_id')
            email_col = get_col(df, 'email', 'e-mail', 'email address', 'email id')
            
            if name_col is None or clicker_col is None:
                return Response(
                    {'error': 'File must have at least "Name" and "Clicker ID" (or "keypad id" / "clicker id") columns.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            reserved = {name_col, clicker_col, email_col} if email_col else {name_col, clicker_col}
            imported = 0
            errors = []
            for idx, row in df.iterrows():
                try:
                    raw_name = row.get(name_col, '')
                    raw_clicker = row.get(clicker_col, '')
                    name = '' if pd.isna(raw_name) else str(raw_name).strip()
                    clicker_id = '' if pd.isna(raw_clicker) else str(raw_clicker).strip()
                    email = None
                    if email_col:
                        raw_email = row.get(email_col, '')
                        email = None if pd.isna(raw_email) else (str(raw_email).strip() or None)
                    extra = {}
                    for col in df.columns:
                        if col not in reserved:
                            val = row.get(col)
                            if pd.notna(val) and val is not None and str(val).strip() not in ('', 'nan'):
                                key = _normalize_col_key(col)
                                if key and key == 'email_id':
                                    if email is None:
                                        email = str(val).strip()
                                elif key:
                                    extra[key] = str(val).strip()
                    
                    if not name or not clicker_id:
                        errors.append(f"Row {idx + 2}: Name and Clicker ID are required.")
                        continue
                    
                    # Avoid UNIQUE constraint on email: if this email is already used by another
                    # participant (different clicker_id), keep it only in extra and set model email to None.
                    if email:
                        existing = Participant.objects.filter(email=email).exclude(clicker_id=clicker_id).first()
                        if existing:
                            extra['email_id'] = extra.get('email_id') or email
                            email = None
                    
                    school_id = get_user_school_id(request.user)
                    participant, created = Participant.objects.update_or_create(
                        clicker_id=clicker_id,
                        created_by=request.user,
                        defaults={'name': name, 'email': email, 'school_id': school_id, 'extra': extra},
                    )
                    
                    if exam_id:
                        try:
                            exam = scope_exams_queryset(Exam.objects.all(), request.user).get(id=exam_id)
                            ExamParticipant.objects.get_or_create(exam=exam, participant=participant)
                        except Exam.DoesNotExist:
                            pass
                    
                    imported += 1
                except Exception as e:
                    errors.append(f"Row {idx + 2}: {str(e)}")
            
            return Response({'imported': imported, 'errors': errors})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def assign_clicker(self, request, pk=None):
        participant = self.get_object()
        clicker_id = request.data.get('clicker_id')
        
        if not clicker_id:
            return Response({'error': 'clicker_id required'}, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = ParticipantSerializer(
            participant,
            data={'clicker_id': str(clicker_id).strip()},
            partial=True,
            context={'request': request},
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


# RBAC: Schools and user management
class SchoolViewSet(viewsets.ModelViewSet):
    """Schools: list/create for Super Admin; list own for School Admin."""
    serializer_class = SchoolSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return scope_schools_queryset(School.objects.all(), self.request.user)

    def perform_create(self, serializer):
        if not can_create_school_admin(self.request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only Super Admin can create schools.')
        serializer.save()


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_school_admin(request):
    """Create a School Admin user. Super Admin only."""
    if not can_create_school_admin(request.user):
        return Response({'error': 'Only Super Admin can create School Admins.'}, status=status.HTTP_403_FORBIDDEN)
    serializer = CreateSchoolAdminSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    try:
        user = serializer.save()
        return Response({
            'id': user.id,
            'email': user.email,
            'role': 'school_admin',
            'school_id': user.profile.school_id,
            'message': 'School Admin created.',
        }, status=status.HTTP_201_CREATED)
    except drf_serializers.ValidationError as e:
        return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def manage_school_admin(request, user_id):
    """Edit or delete a School Admin user. Super Admin only."""
    from django.contrib.auth.models import User
    if not can_create_school_admin(request.user):
        return Response({'error': 'Only Super Admin can manage School Admins.'}, status=status.HTTP_403_FORBIDDEN)
    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return Response({'error': 'User not found.'}, status=status.HTTP_404_NOT_FOUND)
    try:
        profile = user.profile
    except UserProfile.DoesNotExist:
        return Response({'error': 'User has no profile.'}, status=status.HTTP_400_BAD_REQUEST)
    if profile.role != ROLE_SCHOOL_ADMIN:
        return Response({'error': 'User is not a School Admin.'}, status=status.HTTP_400_BAD_REQUEST)

    if request.method == 'DELETE':
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    # PATCH
    data = request.data or {}
    email = data.get('email')
    name = data.get('name')
    school_id = data.get('school_id')
    password = data.get('password')
    from django.contrib.auth.models import User as DjangoUser
    from .models import School

    if email:
        email = email.strip()
        if DjangoUser.objects.exclude(pk=user.pk).filter(email=email).exists():
            return Response({'email': ['A user with this email already exists.']}, status=status.HTTP_400_BAD_REQUEST)
        user.email = email
        user.username = email
    if name is not None:
        user.first_name = (name or '').strip()
    user.save()

    if school_id is not None:
        try:
            school_obj = School.objects.get(pk=school_id)
        except School.DoesNotExist:
            return Response({'school_id': ['School not found.']}, status=status.HTTP_400_BAD_REQUEST)
        profile.school = school_obj
        profile.save()

    if password:
        if len(password) < 8:
            return Response({'password': ['Password must be at least 8 characters.']}, status=status.HTTP_400_BAD_REQUEST)
        user.set_password(password)
        user.save()

    return Response({
        'id': user.id,
        'email': user.email,
        'name': (user.first_name or '').strip() or user.email,
        'school_id': profile.school_id,
        'school_name': profile.school.name if profile.school else '',
    })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_teacher(request):
    """Create a Teacher user. Super Admin or School Admin (for their school)."""
    if not can_create_teacher(request.user):
        return Response({'error': 'Only Super Admin or School Admin can create Teachers.'}, status=status.HTTP_403_FORBIDDEN)
    data = request.data.copy()
    role = get_user_role(request.user)
    if role == ROLE_SCHOOL_ADMIN:
        data['school_id'] = get_user_school_id(request.user)
    serializer = CreateTeacherSerializer(data=data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    try:
        user = serializer.save()
        return Response({
            'id': user.id,
            'email': user.email,
            'role': 'teacher',
            'school_id': user.profile.school_id,
            'message': 'Teacher created.',
        }, status=status.HTTP_201_CREATED)
    except drf_serializers.ValidationError as e:
        return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def manage_teacher(request, user_id):
    """Edit or delete a Teacher user. Super Admin or School Admin (their school)."""
    from django.contrib.auth.models import User
    role = get_user_role(request.user)
    if role not in (ROLE_SUPER_ADMIN, ROLE_SCHOOL_ADMIN):
        return Response({'error': 'Only Super Admin or School Admin can manage Teachers.'}, status=status.HTTP_403_FORBIDDEN)
    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return Response({'error': 'User not found.'}, status=status.HTTP_404_NOT_FOUND)
    try:
        profile = user.profile
    except UserProfile.DoesNotExist:
        return Response({'error': 'User has no profile.'}, status=status.HTTP_400_BAD_REQUEST)
    if profile.role != ROLE_TEACHER:
        return Response({'error': 'User is not a Teacher.'}, status=status.HTTP_400_BAD_REQUEST)

    # School Admin can only manage teachers in their own school
    if role == ROLE_SCHOOL_ADMIN:
        admin_school_id = get_user_school_id(request.user)
        if not admin_school_id or profile.school_id != admin_school_id:
            return Response({'error': 'You can only manage teachers in your school.'}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'DELETE':
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    data = request.data or {}
    email = data.get('email')
    name = data.get('name')
    school_id = data.get('school_id')
    password = data.get('password')
    from django.contrib.auth.models import User as DjangoUser
    from .models import School

    if email:
        email = email.strip()
        if DjangoUser.objects.exclude(pk=user.pk).filter(email=email).exists():
            return Response({'email': ['A user with this email already exists.']}, status=status.HTTP_400_BAD_REQUEST)
        user.email = email
        user.username = email
    if name is not None:
        user.first_name = (name or '').strip()
    user.save()

    if school_id is not None:
        try:
            school_obj = School.objects.get(pk=school_id)
        except School.DoesNotExist:
            return Response({'school_id': ['School not found.']}, status=status.HTTP_400_BAD_REQUEST)
        # For school_admin, ensure they are not moving teacher out of their school
        if role == ROLE_SCHOOL_ADMIN and school_obj.id != admin_school_id:
            return Response({'school_id': ['You can only assign teachers to your school.']}, status=status.HTTP_400_BAD_REQUEST)
        profile.school = school_obj
        profile.save()

    if password:
        if len(password) < 8:
            return Response({'password': ['Password must be at least 8 characters.']}, status=status.HTTP_400_BAD_REQUEST)
        user.set_password(password)
        user.save()

    return Response({
        'id': user.id,
        'email': user.email,
        'name': (user.first_name or '').strip() or user.email,
        'school_id': profile.school_id,
        'school_name': profile.school.name if profile.school else '',
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_exam_owners(request):
    """List School Admins and Teachers for exam owner dropdown. Super Admin sees all; School Admin sees teachers in their school."""
    from django.contrib.auth.models import User
    role = get_user_role(request.user)
    school_id = get_user_school_id(request.user)
    if role == ROLE_TEACHER:
        return Response([])
    qs = UserProfile.objects.filter(role__in=[ROLE_SCHOOL_ADMIN, ROLE_TEACHER]).select_related('user', 'school')
    if role == ROLE_SCHOOL_ADMIN and school_id is not None:
        qs = qs.filter(school_id=school_id)
    out = []
    for p in qs.order_by('school__name', 'user__email'):
        out.append({
            'id': p.user_id,
            'email': p.user.email or '',
            'name': (p.user.first_name or '').strip() or p.user.email or str(p.user_id),
            'role': p.role,
            'school_id': p.school_id,
            'school_name': p.school.name if p.school else '',
        })
    return Response(out)


# Reports Views
# IST timezone for report date/time (dd/mm/yyyy format)
def _report_now_ist():
    """Return current datetime in IST (Asia/Kolkata) for report display."""
    try:
        from zoneinfo import ZoneInfo
    except ImportError:
        from backports.zoneinfo import ZoneInfo
    ist = ZoneInfo('Asia/Kolkata')
    return timezone.now().astimezone(ist)


def _report_datetime_ist():
    """Return current datetime string in IST, format dd/mm/yyyy HH:MM (e.g. 08/03/2026 14:30)."""
    return _report_now_ist().strftime('%d/%m/%Y %H:%M')


def _report_datetime_ist_12h():
    """Return current datetime string in IST with 12-hour AM/PM (e.g. 08/03/2026 02:30 PM)."""
    return _report_now_ist().strftime('%d/%m/%Y %I:%M:%S %p')

# Participant fields to include in report export (label for display, key in Participant.extra or model)
PARTICIPANT_REPORT_FIELDS = [
    ('Name', 'name'),
    ('Clicker ID', 'clicker_id'),
    ('Roll No.', 'roll_no'),
    ('Admission No.', 'admission_no'),
    ('Class', 'class'),
    ('Subject', 'subject'),
    ('Section', 'section'),
    ('Team', 'team'),
    ('Group', 'group'),
    ('House', 'house'),
    ('Gender', 'gender'),
    ('City', 'city'),
    ('UID', 'uid'),
    ('Employee Code', 'employee_code'),
    ('Teacher Name', 'teacher_name'),
    ('Email ID', 'email_id'),
    ('Parent Email ID', 'parent_email_id'),
    ('Parent WhatsApp', 'parent_whatsapp'),
]


def _has_value(v):
    """Return True if v is considered non-empty for report display."""
    if v is None:
        return False
    if isinstance(v, str):
        return bool(v.strip())
    return True


def _participant_detail_columns_with_data(participant_results):
    """Return list of (label, key) from PARTICIPANT_REPORT_FIELDS where at least one result has a non-empty value. Used so we only show user-detail columns that have data."""
    if not participant_results:
        return []
    out = []
    for label, key in PARTICIPANT_REPORT_FIELDS:
        if any(_has_value(r.get(key)) for r in participant_results):
            out.append((label, key))
    return out


def _format_participant_detail_line(user_row, participant_detail_columns):
    """Build client-style single line for participant block: 'Name: x\\nkey: value' for only fields that have data."""
    parts = [f'Name: {user_row.get("name", "") or ""}']
    for _label, key in participant_detail_columns:
        if key in ('name', 'clicker_id'):
            continue
        val = user_row.get(key, '') or ''
        if _has_value(val):
            parts.append(f'{key}: {val}')
    return '\n'.join(parts)


def _participant_report_row(participant):
    """Build a dict of all participant user-data for report export (flat key -> value)."""
    extra = participant.extra or {}
    email_id = extra.get('email_id') or participant.email or ''
    row = {
        'name': participant.name or '',
        'clicker_id': participant.clicker_id or '',
        'email': participant.email or '',
        'roll_no': extra.get('roll_no', ''),
        'admission_no': extra.get('admission_no', ''),
        'class': extra.get('class', ''),
        'subject': extra.get('subject', ''),
        'section': extra.get('section', ''),
        'team': extra.get('team', ''),
        'group': extra.get('group', ''),
        'house': extra.get('house', ''),
        'gender': extra.get('gender', ''),
        'city': extra.get('city', ''),
        'uid': extra.get('uid', ''),
        'employee_code': extra.get('employee_code', ''),
        'teacher_name': extra.get('teacher_name', ''),
        'email_id': email_id,
        'parent_email_id': extra.get('parent_email_id', ''),
        'parent_whatsapp': (
            extra.get('parent_whatsapp')
            or extra.get('parent_phone')
            or extra.get('parent_mobile')
            or ''
        ),
    }
    return row


def _get_exam_report_data(exam):
    """Build report data dict for an exam (no request needed). Used by exam_report and export."""
    attempts = ExamAttempt.objects.filter(exam=exam)
    total_participants = attempts.count()
    if total_participants == 0:
        return {
            'exam_id': exam.id,
            'exam_title': exam.title,
            'total_participants': 0,
            'average_score': 0,
            'highest_score': 0,
            'lowest_score': 0,
            'question_analysis': [],
            'participant_results': []
        }
    avg_score = attempts.aggregate(avg=Avg('score'))['avg'] or 0
    highest = attempts.aggregate(max=Max('score'))['max'] or 0
    lowest = attempts.aggregate(min=Min('score'))['min'] or 0
    questions = exam.exam_questions.select_related('question').order_by('order')
    question_analysis = []
    for eq in questions:
        question = eq.question
        answers = Answer.objects.filter(attempt__exam=exam, question=question)
        total_attempts = answers.count()
        correct_attempts = answers.filter(is_correct=True).count()
        accuracy = (correct_attempts / total_attempts * 100) if total_attempts > 0 else 0
        avg_time = answers.aggregate(avg=Avg('time_taken'))['avg'] or 0
        # Per-option vote counts (0-based index -> count)
        options = question.options or []
        option_votes = [0] * len(options)
        for a in answers:
            sel = a.selected_answer
            if sel is None:
                continue
            if isinstance(sel, list):
                for idx in sel:
                    if isinstance(idx, int) and 0 <= idx < len(option_votes):
                        option_votes[idx] += 1
            elif isinstance(sel, int) and 0 <= sel < len(option_votes):
                option_votes[sel] += 1
        question_analysis.append({
            'question_id': question.id,
            'question_text': question.text,
            'total_attempts': total_attempts,
            'correct_attempts': correct_attempts,
            'accuracy': accuracy,
            'average_time': avg_time,
            'options': options,
            'option_display': getattr(question, 'option_display', 'alpha'),
            'correct_answer': _normalize_correct_answer(question.correct_answer),
            'option_votes': option_votes,
        })
    participant_results = []
    for attempt in attempts.order_by('-score', 'time_taken'):
        p = attempt.participant
        user_row = _participant_report_row(p)
        participant_results.append({
            'participant_id': p.id,
            'participant_name': p.name,
            'name': p.name or '',
            'clicker_id': user_row['clicker_id'],
            'email': user_row['email'],
            'roll_no': user_row['roll_no'],
            'admission_no': user_row['admission_no'],
            'class': user_row['class'],
            'subject': user_row['subject'],
            'section': user_row['section'],
            'team': user_row['team'],
            'group': user_row['group'],
            'house': user_row['house'],
            'gender': user_row['gender'],
            'city': user_row['city'],
            'uid': user_row['uid'],
            'employee_code': user_row['employee_code'],
            'teacher_name': user_row['teacher_name'],
            'email_id': user_row['email_id'],
            'score': float(attempt.score),
            'total_questions': attempt.total_questions,
            'correct_answers': attempt.correct_answers,
            'wrong_answers': attempt.wrong_answers,
            'unattempted': attempt.unattempted,
            'percentage': float(attempt.percentage),
            'rank': 0
        })
    for idx, result in enumerate(participant_results, 1):
        result['rank'] = idx
    report_data = {
        'exam_id': exam.id,
        'exam_title': exam.title,
        'total_participants': total_participants,
        'average_score': float(avg_score),
        'highest_score': float(highest),
        'lowest_score': float(lowest),
        'question_analysis': question_analysis,
        'participant_results': participant_results,
        'participant_detail_columns': _participant_detail_columns_with_data(participant_results),
    }
    return report_data


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exam_report(request, exam_id):
    qs = scope_exams_queryset(Exam.objects.all(), request.user)
    try:
        exam = qs.get(id=exam_id)
    except Exam.DoesNotExist:
        return Response({'error': 'Exam not found'}, status=status.HTTP_404_NOT_FOUND)
    report_data = _get_exam_report_data(exam)
    serializer = ExamReportSerializer(report_data)
    return Response(serializer.data)


def _normalize_correct_answer(correct_answer):
    """Return correct_answer as a list of 0-based indices (for API consistency)."""
    if correct_answer is None:
        return []
    if isinstance(correct_answer, list):
        return [int(x) for x in correct_answer if x is not None]
    return [int(correct_answer)]


def _format_question_options(question):
    """Format question options for display: alpha (A. x, B. y) or numeric (1. x, 2. y)."""
    opts = question.options or []
    display = getattr(question, 'option_display', 'alpha')
    if display == 'numeric':
        return '\n'.join([f'{i + 1}. {opt}' for i, opt in enumerate(opts)])
    # alpha: A, B, C, ...
    return '\n'.join([f'{chr(65 + i)}. {opt}' for i, opt in enumerate(opts)])


def _strip_html(text):
    """Remove HTML tags and unescape entities from text for plain-text report display."""
    if not text or not isinstance(text, str):
        return text or ''
    text = re.sub(r'<[^>]+>', '', text)
    return html_module.unescape(text).strip()


def _option_label_for_index(index, option_display):
    """Return single option label (A/B/C or 1/2/3) for 0-based index."""
    if option_display == 'numeric':
        return str(index + 1)
    return chr(65 + index) if 0 <= index < 26 else str(index + 1)


def _format_attempted_option_label(question, selected_answer):
    """Return the attempted option label(s) for display as numeric (1, 2, 3...). For MCQ/true_false show single number; for multiple_select show comma-separated."""
    if selected_answer is None:
        return ''
    # Single-choice (MCQ, true_false): show one number only (if stored as list, take first)
    qtype = getattr(question, 'type', 'mcq')
    if qtype in ('mcq', 'true_false'):
        if isinstance(selected_answer, list):
            selected_answer = selected_answer[0] if selected_answer else None
        if selected_answer is None:
            return ''
        return str(selected_answer + 1)
    # Multiple-select: show comma-separated numbers
    if isinstance(selected_answer, list):
        return ', '.join(str(i + 1) for i in selected_answer)
    return str(selected_answer + 1)


def _write_results_by_participants_detail_sheet(workbook, exam, participant_detail_columns=None):
    """Write 'Results by Participants(Detail)' sheet to match client format: metadata, then per-participant blocks with Keypad/Name line and table Question|Option|Response|Slide Type|Correct Answer|Speed|Score. Only includes user-detail fields that have data."""
    if participant_detail_columns is None:
        participant_detail_columns = list(PARTICIPANT_REPORT_FIELDS)
    sheet_name = 'Results by Participants(Detail)'
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name, 0)
    red_header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)

    exam_questions = list(exam.exam_questions.select_related('question').order_by('order'))
    attempts = list(ExamAttempt.objects.filter(exam=exam).order_by('participant__clicker_id').select_related('participant'))
    answer_map = {}
    for a in Answer.objects.filter(attempt__exam=exam).select_related('attempt', 'question'):
        answer_map[(a.attempt_id, a.question_id)] = a

    # Rank by score (desc), then time_taken (asc: faster = better). Rank 1 = best.
    attempts_ranked = sorted(
        ExamAttempt.objects.filter(exam=exam).values_list('id', 'score', 'time_taken'),
        key=lambda x: (-float(x[1]), x[2] or 0)
    )
    rank_by_attempt_id = {aid: rank for rank, (aid, _s, _t) in enumerate(attempts_ranked, 1)}
    total_participants = len(attempts_ranked)

    row_num = 1
    ws.cell(row=row_num, column=1, value='Results by Participants(Detail)')
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
    row_num += 1
    ws.cell(row=row_num, column=1, value=_report_datetime_ist())
    row_num += 1
    ws.cell(row=row_num, column=1, value=f'Questions Count: {len(exam_questions)}')
    row_num += 2
    # Metadata block (client format: label in A, value in B; Theme/Faculty/Batch empty)
    for label, col_b in [
        ('Sl No.', ''),
        ('Exam Unique Id', str(exam.id)),
        ('Exam Name', exam.title or ''),
        ('Theme', ''),
        ('Faculty', ''),
        ('Subject', exam.title or ''),
        ('Batch', ''),
    ]:
        ws.cell(row=row_num, column=1, value=label)
        ws.cell(row=row_num, column=2, value=col_b)
        row_num += 1
    row_num += 1

    for attempt in attempts:
        participant = attempt.participant
        keypad = getattr(participant, 'clicker_id', None) or participant.name or str(participant.id)
        user_row = _participant_report_row(participant)
        # Client format: row 1 = Keypad No. X, Correct Count, Score; row 2 = Name + details (only fields with data), Correct Rate %, Ranking
        ws.cell(row=row_num, column=1, value=f'Keypad No. {keypad}')
        ws.cell(row=row_num, column=2, value=f'Correct Count: {attempt.correct_answers}')
        ws.cell(row=row_num, column=3, value=f'Score: {int(attempt.score)}')
        row_num += 1
        detail_line = _format_participant_detail_line(user_row, participant_detail_columns)
        correct_rate = (attempt.correct_answers / attempt.total_questions * 100) if attempt.total_questions else 0
        ws.cell(row=row_num, column=1, value=detail_line)
        ws.cell(row=row_num, column=2, value=f'Correct Rate: {correct_rate:.2f}%')
        ws.cell(row=row_num, column=3, value=f'Ranking: {rank_by_attempt_id.get(attempt.id, "")}/{total_participants}')
        row_num += 1
        # Table header: Question, Option, Response, Slide Type, Correct Answer, Speed, Score
        headers = ['Question', 'Option', 'Response', 'Slide Type', 'Correct Answer', 'Speed', 'Score']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=col, value=h)
            c.fill = red_header_fill
            c.font = white_font
        row_num += 1
        for order_idx, eq in enumerate(exam_questions, start=1):
            q = eq.question
            options_text = _format_question_options(q)
            answer = answer_map.get((attempt.id, q.id))
            correct_idx = q.correct_answer
            if isinstance(correct_idx, list):
                correct_display = correct_idx[0] + 1 if correct_idx else ''
            else:
                correct_display = int(correct_idx) + 1 if correct_idx is not None else ''
            if answer:
                sel = answer.selected_answer
                if isinstance(sel, list):
                    response_val = (sel[0] + 1) if sel else ''
                else:
                    response_val = int(sel) + 1 if sel is not None else ''
                # Speed = time taken for this question in seconds (integer from DB)
                speed = answer.time_taken if answer.time_taken is not None else 0
                score = int(eq.positive_marks) if answer.is_correct else -int(eq.negative_marks)
            else:
                response_val = ''
                speed = ''
                score = 0
            ws.cell(row=row_num, column=1, value=f'{order_idx}. {_strip_html(q.text)}')
            ws.cell(row=row_num, column=2, value=options_text)
            ws.cell(row=row_num, column=3, value=response_val)
            ws.cell(row=row_num, column=4, value='Choice')
            ws.cell(row=row_num, column=5, value=correct_display)
            ws.cell(row=row_num, column=6, value=speed)
            ws.cell(row=row_num, column=7, value=score)
            row_num += 1
        row_num += 1

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 8


def _sanitize_sheet_name(name):
    """Excel sheet names: max 31 chars, no \\ / * ? : [ ]"""
    for c in '\\/*?:[]':
        name = name.replace(c, '_')
    return (name or 'Sheet')[:31]


def _write_results_by_participants_individual(workbook, exam, participant_detail_columns=None):
    """One sheet per participant (sheet name = keypad id). Same layout as reference: title, date, questions count, metadata, Keypad summary, then Question|Option|Response|Slide Type|Correct Answer|Speed|Score. Only includes user-detail fields that have data (if participant_detail_columns provided)."""
    if participant_detail_columns is None:
        participant_detail_columns = list(PARTICIPANT_REPORT_FIELDS)
    red_header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)
    bold_font = Font(bold=True)

    exam_questions = list(exam.exam_questions.select_related('question').order_by('order'))
    attempts = list(ExamAttempt.objects.filter(exam=exam).order_by('participant__clicker_id').select_related('participant'))
    answer_map = {}
    for a in Answer.objects.filter(attempt__exam=exam).select_related('attempt', 'question'):
        answer_map[(a.attempt_id, a.question_id)] = a

    # Rank by score (desc), then time_taken (asc). Rank 1 = best.
    attempts_ranked = sorted(
        ExamAttempt.objects.filter(exam=exam).values_list('id', 'score', 'time_taken'),
        key=lambda x: (-float(x[1]), x[2] or 0)
    )
    rank_by_attempt_id = {aid: rank for rank, (aid, _s, _t) in enumerate(attempts_ranked, 1)}
    total_participants = len(attempts_ranked)

    n_questions = len(exam_questions)
    report_date = _report_datetime_ist()

    for attempt in attempts:
        participant = attempt.participant
        keypad = getattr(participant, 'clicker_id', None) or str(participant.id)
        user_row = _participant_report_row(participant)
        sheet_name = _sanitize_sheet_name(str(keypad))
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        ws = workbook.create_sheet(sheet_name)

        row = 1
        ws.cell(row=row, column=1, value='Results by Participants(Detail)')
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1
        ws.cell(row=row, column=1, value=report_date)
        row += 1
        ws.cell(row=row, column=1, value=f'Questions Count: {n_questions}')
        row += 2
        # Metadata block (client format)
        for label, col_b in [
            ('Sl No.', ''),
            ('Exam Unique Id', str(exam.id)),
            ('Exam Name', exam.title or ''),
            ('Theme', ''),
            ('Faculty', ''),
            ('Subject', exam.title or ''),
            ('Batch', ''),
        ]:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=col_b)
            row += 1
        row += 1
        # Keypad summary row + Name/details row (only fields with data)
        ws.cell(row=row, column=1, value=f'Keypad No. {keypad}')
        ws.cell(row=row, column=2, value=f'Correct Count: {attempt.correct_answers}')
        ws.cell(row=row, column=3, value=f'Score: {int(attempt.score)}')
        row += 1
        detail_line = _format_participant_detail_line(user_row, participant_detail_columns)
        correct_rate = (attempt.correct_answers / attempt.total_questions * 100) if attempt.total_questions else 0
        ws.cell(row=row, column=1, value=detail_line)
        ws.cell(row=row, column=2, value=f'Correct Rate: {correct_rate:.2f}%')
        ws.cell(row=row, column=3, value=f'Ranking: {rank_by_attempt_id.get(attempt.id, "")}/{total_participants}')
        row += 1
        # Table header
        headers = ['Question', 'Option', 'Response', 'Slide Type', 'Correct Answer', 'Speed', 'Score']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = red_header_fill
            c.font = white_font
        row += 1
        # One row per question (match reference: Response = 1-based index, Slide Type = 'Choice', Correct Answer = 1-based correct index, Speed = decimal)
        for order_idx, eq in enumerate(exam_questions, start=1):
            q = eq.question
            options_text = _format_question_options(q)
            answer = answer_map.get((attempt.id, q.id))
            correct_idx = q.correct_answer
            if isinstance(correct_idx, list):
                correct_display = correct_idx[0] + 1 if correct_idx else ''
            else:
                correct_display = int(correct_idx) + 1 if correct_idx is not None else ''
            if answer:
                sel = answer.selected_answer
                if isinstance(sel, list):
                    response_val = (sel[0] + 1) if sel else ''
                else:
                    response_val = int(sel) + 1 if sel is not None else ''
                # Speed = time taken for this question in seconds
                speed = answer.time_taken if answer.time_taken is not None else 0
                score = int(eq.positive_marks) if answer.is_correct else -int(eq.negative_marks)
            else:
                response_val = ''
                speed = ''
                score = 0
            ws.cell(row=row, column=1, value=f'{order_idx}. {_strip_html(q.text)}')
            ws.cell(row=row, column=2, value=options_text)
            ws.cell(row=row, column=3, value=response_val)
            ws.cell(row=row, column=4, value='Choice')
            ws.cell(row=row, column=5, value=correct_display)
            ws.cell(row=row, column=6, value=speed)
            ws.cell(row=row, column=7, value=score)
            row += 1

        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 8

    # Remove default empty sheet if present (we didn't write any pandas sheet first)
    for default in ('Sheet', 'Sheet1'):
        if default in workbook.sheetnames and len(workbook.sheetnames) > 1:
            del workbook[default]
            break


def _write_results_by_questions_sheet(workbook, exam):
    """One sheet 'Results by Questions': per-question blocks with Option, Voted, Percentage (like reference)."""
    sheet_name = 'Results by Questions'
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name, 0)
    red_header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

    exam_questions = list(exam.exam_questions.select_related('question').order_by('order'))
    answers_by_question = {}  # question_id -> list of selected_answer (int or list)
    for a in Answer.objects.filter(attempt__exam=exam).values_list('question_id', 'selected_answer'):
        qid, sel = a[0], a[1]
        answers_by_question.setdefault(qid, []).append(sel)


    row = 1
    ws.cell(row=row, column=1, value='Results by Questions')
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value=exam.title or '')
    row += 1
    ws.cell(row=row, column=1, value=_report_datetime_ist_12h())
    row += 2
    for label, col_b in [
        ('Sl No.', ''),
        ('Exam Unique Id', str(exam.id)),
        ('Exam Name', exam.title or ''),
        ('Theme', ''),
        ('Faculty', ''),
        ('Subject', exam.title or ''),
        ('Batch', ''),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=col_b)
        row += 1
    row += 1

    for order_idx, eq in enumerate(exam_questions, start=1):
        q = eq.question
        options = q.options or []
        correct_rate = 0.0
        total_answers = 0
        option_votes = [0] * len(options)  # vote count per option index

        selections = answers_by_question.get(q.id) or []
        total_answers = len(selections)
        for sel in selections:
            if isinstance(sel, list):
                for i in sel:
                    if 0 <= i < len(option_votes):
                        option_votes[i] += 1
            else:
                if 0 <= sel < len(option_votes):
                    option_votes[sel] += 1
        if total_answers > 0:
            correct_count = sum(1 for s in selections if _answer_is_correct(s, q.correct_answer))
            correct_rate = round(correct_count / total_answers * 100, 2)

        ws.cell(row=row, column=1, value=f'{order_idx}. {_strip_html(q.text)}')
        ws.cell(row=row, column=2, value='Slide Type: Choice')
        ws.cell(row=row, column=3, value=f'Correct Rate: {correct_rate:.2f}%')
        row += 1
        ws.cell(row=row, column=1, value='Option')
        ws.cell(row=row, column=2, value='Voted')
        ws.cell(row=row, column=3, value='Percentage')
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = red_header_fill
            ws.cell(row=row, column=c).font = white_font
        row += 1

        correct_indices = set(_normalize_correct_answer(q.correct_answer))
        for i, opt in enumerate(options):
            letter = chr(ord('A') + i) if i < 26 else 'X'
            opt_label = f'{i+1}/{letter}. {opt}'
            vote_count = option_votes[i] if i < len(option_votes) else 0
            pct = (vote_count / total_answers * 100) if total_answers else 0
            is_correct = i in correct_indices
            ws.cell(row=row, column=1, value=opt_label)
            ws.cell(row=row, column=2, value=vote_count)
            ws.cell(row=row, column=3, value=f'{pct:.2f}%')
            if is_correct:
                for c in range(1, 4):
                    ws.cell(row=row, column=c).fill = green_fill
            row += 1
        ws.cell(row=row, column=1, value='Voted')
        ws.cell(row=row, column=2, value=total_answers)
        ws.cell(row=row, column=3, value='100.00%')
        row += 2

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14


# Personal Achievement and Detail: one sheet with metadata + table of all assigned participants with keypad, name, roll_no, ..., Score, Correct Rate, Ranking, and per-question columns (1-S1, 2-S2, ...)
PERSONAL_ACHIEVEMENT_HEADERS = [
    'Keypad No.', 'Name', 'roll number', 'admission number', 'class', 'subject', 'section',
    'team', 'group', 'house', 'gender', 'city', 'uid', 'employee code', 'teaccher name', 'Email ID',
    'Score', 'Correct Rate', 'Ranking',
]


def _write_personal_achievement_and_detail_sheet(workbook, exam):
    """Write 'Personal Achievement and Detail' sheet: title, Voted, exam info, then table of participants with details + score + per-question columns."""
    sheet_name = 'Personal Achievement and Detail'
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name, 0)

    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # light green for correct
    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')    # light red for wrong

    exam_questions = list(exam.exam_questions.select_related('question').order_by('order'))
    assigned = list(ExamParticipant.objects.filter(exam=exam).select_related('participant').order_by('participant__clicker_id'))
    # Natural order by keypad id (1, 2, ... 9, 10, 11)
    assigned.sort(key=lambda ep: (len(ep.participant.clicker_id or ''), ep.participant.clicker_id or ''))
    attempts_by_participant = {a.participant_id: a for a in ExamAttempt.objects.filter(exam=exam).select_related('participant')}
    answer_map = {}  # (attempt_id, question_id) -> Answer
    for a in Answer.objects.filter(attempt__exam=exam).select_related('attempt', 'question'):
        answer_map[(a.attempt_id, a.question_id)] = a

    # Rank attempted by score (desc), then time (asc)
    attempted_list = sorted(
        [attempts_by_participant[ep.participant_id] for ep in assigned if ep.participant_id in attempts_by_participant],
        key=lambda a: (-float(a.score), a.time_taken)
    )
    rank_by_attempt = {a.id: idx + 1 for idx, a in enumerate(attempted_list)}
    voted_count = len(attempted_list)

    assigned_sorted = sorted(assigned, key=lambda ep: ep.participant.clicker_id or '')

    row = 1
    ws.cell(row=row, column=1, value='Personal Achievement and Detail')
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value=f'Voted:{voted_count}')
    row += 1
    ws.cell(row=row, column=1, value=exam.title or '')
    row += 1
    ws.cell(row=row, column=1, value=_report_datetime_ist())
    row += 1
    for label, col_b in [
        ('Sl No.', ''),
        ('Exam Unique Id', str(exam.id)),
        ('Exam Name', exam.title or ''),
        ('Theme', ''),
        ('Faculty', ''),
        ('Subject', exam.title or ''),
        ('Batch', ''),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=col_b)
        row += 1
    # Table header row: Keypad No., Name, then only detail columns that have data (excl. name/clicker_id), then Score, Correct Rate, Ranking, then question columns
    user_rows = [_participant_report_row(ep.participant) for ep in assigned_sorted]
    detail_columns = _participant_detail_columns_with_data(user_rows)
    detail_columns_no_id = [(label, key) for (label, key) in detail_columns if key not in ('name', 'clicker_id')]
    headers = ['Keypad No.', 'Name'] + [label for (label, _) in detail_columns_no_id] + ['Score', 'Correct Rate', 'Ranking']
    for order_idx, eq in enumerate(exam_questions, start=1):
        headers.append(f'{order_idx}-S{order_idx}')
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    row += 1

    for ep in assigned_sorted:
        p = ep.participant
        user_row = _participant_report_row(p)
        attempt = attempts_by_participant.get(p.id)
        if attempt:
            score_val = float(attempt.score)
            correct_rate = (attempt.correct_answers / attempt.total_questions * 100) if attempt.total_questions else 0
            rank_val = rank_by_attempt.get(attempt.id, '')
        else:
            score_val = 'ABSENT'
            correct_rate = 0
            rank_val = ''

        col = 1
        ws.cell(row=row, column=col, value=p.clicker_id or '')
        col += 1
        ws.cell(row=row, column=col, value=p.name or '')
        col += 1
        for _label, key in detail_columns_no_id:
            ws.cell(row=row, column=col, value=user_row.get(key, '') or '')
            col += 1
        ws.cell(row=row, column=col, value=score_val)
        col += 1
        ws.cell(row=row, column=col, value=f'{correct_rate:.2f}%' if attempt else '0.00%')
        col += 1
        ws.cell(row=row, column=col, value=rank_val)
        col += 1
        for eq in exam_questions:
            if attempt:
                ans = answer_map.get((attempt.id, eq.question_id))
                if ans:
                    attempted_label = _format_attempted_option_label(eq.question, ans.selected_answer)
                    if ans.is_correct:
                        cell = ws.cell(row=row, column=col, value=attempted_label)
                        cell.fill = green_fill
                    else:
                        cell = ws.cell(row=row, column=col, value=attempted_label)
                        cell.fill = red_fill
                else:
                    ws.cell(row=row, column=col, value='')
                col += 1
            else:
                ws.cell(row=row, column=col, value='')
                col += 1
        row += 1

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    # Remove default empty sheet if present
    for default in ('Sheet', 'Sheet1'):
        if default in workbook.sheetnames and len(workbook.sheetnames) > 1:
            del workbook[default]
            break


def _answer_is_correct(selected, correct_answer):
    """Return True if selected answer matches correct_answer (int or list).
    selected_answer in DB is often stored as list for MCQ (e.g. [1]); correct_answer may be int or list.
    """
    if selected is None or correct_answer is None:
        return False
    # Normalize when correct_answer is list (multiple correct)
    if isinstance(correct_answer, list):
        sel_list = selected if isinstance(selected, list) else [selected]
        return set(sel_list) == set(correct_answer)
    # MCQ: correct_answer is int; selected may be int or single-element list from DB
    if isinstance(selected, list):
        return len(selected) == 1 and selected[0] == correct_answer
    return selected == correct_answer


def _build_export_http_response(request, exam):
    """Build Excel or CSV download for exam report. Query param: format=excel|csv (case-insensitive, default excel). layout=personal_achievement for Personal Achievement sheet only."""
    raw = (request.query_params.get('format') or 'excel').strip().lower()
    format_type = 'excel' if raw in ('excel', 'xlsx', '') else 'csv'
    layout = (request.query_params.get('layout') or '').strip().lower()
    report_data = _get_exam_report_data(exam)
    detail_columns = report_data.get('participant_detail_columns') or []
    detail_keys = [k for (_, k) in detail_columns]
    base_columns = ['participant_id', 'score', 'total_questions', 'correct_answers', 'wrong_answers', 'unattempted', 'percentage', 'rank']
    export_columns = ['participant_id'] + detail_keys + base_columns
    df = pd.DataFrame(report_data['participant_results'])
    df = df[[c for c in export_columns if c in df.columns]]
    # Ensure ranking order (rank 1 first)
    if 'rank' in df.columns:
        df = df.sort_values('rank', ascending=True)
    output = BytesIO()
    if format_type == 'excel':
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if layout == 'personal_achievement':
                _write_personal_achievement_and_detail_sheet(writer.book, exam)
                filename = f'report-{exam.id}-personal-achievement.xlsx'
            else:
                _write_results_by_participants_detail_sheet(writer.book, exam, participant_detail_columns=detail_columns)
                filename = f'report-{exam.id}.xlsx'
    else:
        content_type = 'text/csv'
        filename = f'report-{exam.id}.csv'
        df.to_csv(output, index=False)
    output.seek(0)
    return Response(
        output.getvalue(),
        content_type=content_type,
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_report(request, exam_id):
    """Export report: GET /api/report-export/<exam_id>/?format=excel|csv. Returns file download as DRF Response."""
    try:
        exam = Exam.objects.get(id=exam_id)
    except Exam.DoesNotExist:
        return Response(
            {'error': 'Exam not found', 'exam_id': exam_id, 'message': 'No exam with this id. Run: python manage.py seed_dummy_data'},
            status=status.HTTP_404_NOT_FOUND
        )
    return _build_export_http_response(request, exam)


# Use binary renderer so file bytes are not JSON-encoded
export_report.renderer_classes = [BinaryFileRenderer]


def _build_export_file_response(format_type, exam, layout=None):
    """Build Excel or CSV as Django HttpResponse. layout='individual' => one sheet per participant; layout='questions' => Results by Questions; layout='personal_achievement' => Personal Achievement and Detail."""
    report_data = _get_exam_report_data(exam)
    detail_columns = report_data.get('participant_detail_columns') or []
    detail_keys = [k for (_, k) in detail_columns]
    base_columns = ['participant_id', 'score', 'total_questions', 'correct_answers', 'wrong_answers', 'unattempted', 'percentage', 'rank']
    export_columns = ['participant_id'] + detail_keys + base_columns
    df = pd.DataFrame(report_data['participant_results'])
    df = df[[c for c in export_columns if c in df.columns]]
    # Ensure ranking order (rank 1 first)
    if 'rank' in df.columns:
        df = df.sort_values('rank', ascending=True)
    output = BytesIO()
    if format_type == 'excel':
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            layout_l = (layout or '').strip().lower()
            if layout_l == 'individual':
                _write_results_by_participants_individual(writer.book, exam, participant_detail_columns=detail_columns)
                filename = f'report-{exam.id}-individual.xlsx'
            elif layout_l == 'questions':
                _write_results_by_questions_sheet(writer.book, exam)
                # Remove default sheet if pandas created one
                for default in ('Sheet', 'Sheet1'):
                    if default in writer.book.sheetnames and len(writer.book.sheetnames) > 1:
                        del writer.book[default]
                        break
                filename = f'report-{exam.id}-questions.xlsx'
            elif layout_l == 'personal_achievement':
                _write_personal_achievement_and_detail_sheet(writer.book, exam)
                filename = f'report-{exam.id}-personal-achievement.xlsx'
            else:
                _write_results_by_participants_detail_sheet(writer.book, exam, participant_detail_columns=detail_columns)
                filename = f'report-{exam.id}.xlsx'
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    else:
        df.to_csv(output, index=False)
        content_type = 'text/csv'
        filename = f'report-{exam.id}.csv'
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@require_GET
def export_report_http(request, exam_id):
    """Plain Django view for report export. Uses HttpRequest and returns HttpResponse to avoid DRF assertion."""
    auth = JWTAuthentication()
    try:
        user_auth = auth.authenticate(request)
    except Exception:
        user_auth = None
    if user_auth is None:
        return JsonResponse({'detail': 'Authentication credentials were not provided.'}, status=401)
    # JWTAuthentication.authenticate() returns a (user, token) tuple, but
    # since this is a plain Django view we must attach it to request manually.
    user, token = user_auth
    request.user = user
    request.auth = token
    exam_qs = scope_exams_queryset(Exam.objects.all(), request.user)
    try:
        exam = exam_qs.get(id=exam_id)
    except Exam.DoesNotExist:
        return JsonResponse(
            {'error': 'Exam not found', 'exam_id': exam_id, 'message': 'No exam with this id. Run: python manage.py seed_dummy_data'},
            status=404
        )
    raw = (request.GET.get('format') or 'excel').strip().lower()
    format_type = 'excel' if raw in ('excel', 'xlsx', '') else 'csv'
    layout = (request.GET.get('layout') or '').strip().lower()  # 'individual' => one sheet per participant
    return _build_export_file_response(format_type, exam, layout=layout)


# Dashboard Views
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard(request):
    user = request.user
    exams = scope_exams_queryset(Exam.objects.all(), user)
    
    # Stats
    total_exams = exams.count()
    participants_qs = scope_participants_queryset(Participant.objects.all(), user)
    total_participants = participants_qs.count()
    
    # Calculate average score from all attempts (only for user's visible exams)
    all_attempts = ExamAttempt.objects.filter(exam__in=exams).select_related('exam')
    if all_attempts.exists():
        total_percentage = 0
        count = 0
        for attempt in all_attempts:
            if attempt.total_questions > 0:
                percentage = (float(attempt.score) / (attempt.total_questions * float(attempt.exam.positive_marking))) * 100
                total_percentage += percentage
                count += 1
        avg_score = total_percentage / count if count > 0 else 0
    else:
        avg_score = 0
    
    # Attendance rate (participants who attempted / total participants)
    total_assigned = ExamParticipant.objects.filter(exam__in=exams).count()
    attempted = all_attempts.count()
    attendance_rate = (attempted / total_assigned * 100) if total_assigned > 0 else 0
    
    stats = {
        'total_exams': total_exams,
        'total_participants': total_participants,
        'average_score': float(avg_score),
        'attendance_rate': float(attendance_rate)
    }
    
    # Recent exams
    recent_exams = []
    for exam in exams[:5]:
        attempts = ExamAttempt.objects.filter(exam=exam).select_related('exam')
        participant_count = attempts.count()
        if attempts.exists():
            total_percentage = 0
            count = 0
            for attempt in attempts:
                if attempt.total_questions > 0:
                    percentage = (float(attempt.score) / (attempt.total_questions * float(exam.positive_marking))) * 100
                    total_percentage += percentage
                    count += 1
            avg_exam_score = total_percentage / count if count > 0 else 0
        else:
            avg_exam_score = 0
        
        recent_exams.append({
            'id': exam.id,
            'title': exam.title,
            'created_at': exam.created_at,
            'participant_count': participant_count,
            'average_score': float(avg_exam_score)
        })
    
    # Performance data (last 7 days)
    performance_data = []
    for i in range(6, -1, -1):
        date = timezone.now() - timedelta(days=i)
        date_str = date.strftime('%Y-%m-%d')
        
        day_attempts = ExamAttempt.objects.filter(
            exam__created_by=user,
            submitted_at__date=date.date()
        ).select_related('exam')
        
        if day_attempts.exists():
            total_percentage = 0
            count = 0
            for attempt in day_attempts:
                if attempt.total_questions > 0:
                    percentage = (float(attempt.score) / (attempt.total_questions * float(attempt.exam.positive_marking))) * 100
                    total_percentage += percentage
                    count += 1
            avg_day_score = total_percentage / count if count > 0 else 0
        else:
            avg_day_score = 0
        participant_count = day_attempts.values('participant').distinct().count()
        
        performance_data.append({
            'date': date_str,
            'score': float(avg_day_score),
            'participants': participant_count
        })
    
    dashboard_data = {
        'stats': stats,
        'recent_exams': recent_exams,
        'performance_data': performance_data
    }
    
    serializer = DashboardDataSerializer(dashboard_data)
    return Response(serializer.data)


def _build_student_performance_rows(user, query_params):
    """Build role-scoped student performance rows from query params."""
    # Role-wise scoping
    participants_qs = scope_participants_queryset(Participant.objects.all(), user)
    exams_qs = scope_exams_queryset(Exam.objects.all(), user)

    admission_no = (query_params.get('admission_no') or '').strip().lower()
    roll_no = (query_params.get('roll_no') or '').strip().lower()
    student_name = (query_params.get('student_name') or '').strip().lower()
    class_name = (query_params.get('class_name') or '').strip().lower()
    section = (query_params.get('section') or '').strip().lower()
    teacher_name = (query_params.get('teacher_name') or '').strip().lower()
    subject = (query_params.get('subject') or '').strip().lower()
    from_date = (query_params.get('from_date') or '').strip()
    to_date = (query_params.get('to_date') or '').strip()

    attempts_qs = ExamAttempt.objects.filter(
        exam__in=exams_qs,
        participant__in=participants_qs,
    ).select_related('participant', 'exam')

    if from_date:
        attempts_qs = attempts_qs.filter(submitted_at__date__gte=from_date)
    if to_date:
        attempts_qs = attempts_qs.filter(submitted_at__date__lte=to_date)

    rows_by_participant = {}
    for attempt in attempts_qs:
        p = attempt.participant
        if not p:
            continue
        extra = p.extra or {}
        name_val = (p.name or '').strip()
        row = rows_by_participant.get(p.id)
        if row is None:
            row = {
                'participant_id': p.id,
                'admission_no': (extra.get('admission_no') or '').strip() if isinstance(extra, dict) else '',
                'roll_no': (extra.get('roll_no') or '').strip() if isinstance(extra, dict) else '',
                'student_name': name_val,
                'class_name': (extra.get('class') or '').strip() if isinstance(extra, dict) else '',
                'section': (extra.get('section') or '').strip() if isinstance(extra, dict) else '',
                'teacher_name': (extra.get('teacher_name') or '').strip() if isinstance(extra, dict) else '',
                'subject': (extra.get('subject') or '').strip() if isinstance(extra, dict) else '',
                '_attempts': 0,
                '_pct_sum': 0.0,
            }
            rows_by_participant[p.id] = row

        pct = 0.0
        if attempt.total_questions and float(getattr(attempt.exam, 'positive_marking', 1) or 1) > 0:
            pct = (float(attempt.score) / (attempt.total_questions * float(attempt.exam.positive_marking))) * 100
        row['_attempts'] += 1
        row['_pct_sum'] += pct

    rows = []
    for row in rows_by_participant.values():
        # Apply text filters on flattened row values
        if admission_no and admission_no not in (row['admission_no'] or '').lower():
            continue
        if roll_no and roll_no not in (row['roll_no'] or '').lower():
            continue
        if student_name and student_name not in (row['student_name'] or '').lower():
            continue
        if class_name and class_name not in (row['class_name'] or '').lower():
            continue
        if section and section not in (row['section'] or '').lower():
            continue
        if teacher_name and teacher_name not in (row['teacher_name'] or '').lower():
            continue
        if subject and subject not in (row['subject'] or '').lower():
            continue

        attempts_count = row.pop('_attempts', 0) or 0
        pct_sum = row.pop('_pct_sum', 0.0) or 0.0
        row['total_percentage'] = round((pct_sum / attempts_count), 2) if attempts_count > 0 else 0.0
        rows.append(row)

    rows.sort(key=lambda x: (-float(x.get('total_percentage') or 0), (x.get('student_name') or '').lower()))
    return rows


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_performance_report(request):
    """
    GET /api/reports/student-performance/
    Role-scoped student performance rows with optional filters:
      admission_no, roll_no, student_name, class_name, section, teacher_name, subject,
      from_date (YYYY-MM-DD), to_date (YYYY-MM-DD)
    """
    rows = _build_student_performance_rows(request.user, request.query_params)
    return Response({'count': len(rows), 'results': rows})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_performance_report_export(request):
    """
    GET /api/reports/student-performance/export/?file_format=excel|csv&...
    Exports student performance report with current filters.
    """
    rows = _build_student_performance_rows(request.user, request.query_params)
    # DRF reserves `format` for content negotiation in some setups; prefer `file_format`.
    raw = (request.query_params.get('file_format') or request.query_params.get('format') or 'excel').strip().lower()
    format_type = 'excel' if raw in ('excel', 'xlsx', '') else 'csv'

    columns = [
        'admission_no', 'roll_no', 'student_name', 'class_name',
        'section', 'teacher_name', 'subject', 'total_percentage'
    ]
    df = pd.DataFrame(rows, columns=columns)
    df = df.rename(columns={
        'admission_no': 'Admission No',
        'roll_no': 'Roll No',
        'student_name': 'Student Name',
        'class_name': 'Class',
        'section': 'Section',
        'teacher_name': 'Teacher Name',
        'subject': 'Subject',
        'total_percentage': 'Total Percentage %',
    })

    output = BytesIO()
    if format_type == 'excel':
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Student Performance', index=False)
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = 'student-performance-report.xlsx'
    else:
        df.to_csv(output, index=False)
        content_type = 'text/csv'
        filename = 'student-performance-report.csv'

    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# Leaderboard Views
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def leaderboard(request, exam_id):
    qs = scope_exams_queryset(Exam.objects.all(), request.user)
    try:
        exam = qs.get(id=exam_id)
    except Exam.DoesNotExist:
        return Response({'error': 'Exam not found'}, status=status.HTTP_404_NOT_FOUND)
    
    attempts = ExamAttempt.objects.filter(exam=exam).order_by('-score', 'time_taken')
    
    entries = []
    for rank, attempt in enumerate(attempts, 1):
        entries.append({
            'rank': rank,
            'participant_id': attempt.participant.id,
            'participant_name': attempt.participant.name,
            'score': float(attempt.score),
            'percentage': float(attempt.percentage),
            'total_questions': attempt.total_questions,
            'correct_answers': attempt.correct_answers,
            'time_taken': attempt.time_taken
        })
    
    leaderboard_data = {
        'exam_id': exam.id,
        'exam_title': exam.title,
        'entries': entries,
        'generated_at': timezone.now()
    }
    
    serializer = LeaderboardSerializer(leaderboard_data)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_leaderboard(request, exam_id):
    """
    GET /api/leaderboard/exams/{exam_id}/export/?format=excel|pdf
    Downloads leaderboard report as Excel or PDF.
    """
    # DRF uses `?format=` for content negotiation; use `file_format` instead.
    raw = (request.GET.get('file_format') or request.GET.get('format') or 'excel').strip().lower()
    format_type = 'pdf' if raw in ('pdf',) else 'excel'

    qs = scope_exams_queryset(Exam.objects.all(), request.user)
    try:
        exam = qs.get(id=exam_id)
    except Exam.DoesNotExist:
        return Response({'error': 'Exam not found'}, status=status.HTTP_404_NOT_FOUND)

    attempts = ExamAttempt.objects.filter(exam=exam).order_by('-score', 'time_taken')
    entries = []
    for rank, attempt in enumerate(attempts, 1):
        entries.append({
            'Rank': rank,
            'Participant ID': attempt.participant.id,
            'Participant': attempt.participant.name,
            'Score': float(attempt.score),
            'Correct': attempt.correct_answers,
            'Total Questions': attempt.total_questions,
            'Percentage': float(attempt.percentage),
            'Time Taken (sec)': attempt.time_taken,
        })

    filename_base = f'leaderboard-{exam.id}'
    if format_type == 'excel':
        output = BytesIO()
        df = pd.DataFrame(entries, columns=[
            'Rank', 'Participant ID', 'Participant', 'Score', 'Correct',
            'Total Questions', 'Percentage', 'Time Taken (sec)'
        ])
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Leaderboard', index=False)
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
        return response

    # PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f'Leaderboard - {exam.title}', styles['Title']),
        Spacer(1, 12),
        Paragraph(f'Generated at: {timezone.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']),
        Spacer(1, 12),
    ]

    table_data = [[
        'Rank', 'Participant', 'Score', 'Correct', 'Total', 'Percentage', 'Time Taken (sec)'
    ]] + [
        [e['Rank'], e['Participant'], e['Score'], e['Correct'], e['Total Questions'], e['Percentage'], e['Time Taken (sec)']]
        for e in entries
    ]
    table = Table(table_data, repeatRows=1, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
    return response
